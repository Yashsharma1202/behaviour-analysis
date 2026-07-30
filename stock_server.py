"""
stock_server.py
===============================================================================
NSE STOCK BROWSER — web version, hosted on localhost.
-------------------------------------------------------------------------------
"One stock at a time, start from the first", served as a web page.

TABS per stock:
    Fundamentals      — snapshot cards + annual & quarterly charts (all stocks)
    Announcements     — NSE announcement feed        }
    Board Meetings    — board-meeting intimations     }  read from a per-stock
    Corporate Actions — dividends / bonus / splits …   }  <SYMBOL>/ folder
    Financial Results — quarterly / annual result rows}

The event tabs read <SYMBOL>/announcements.csv, board_meetings.csv,
corporate_actions.csv and financial_results.csv. Right now only the RELIANCE/
folder has these, so RELIANCE shows full detail and other stocks show a clear
"no event feed yet" note — the moment you drop those four CSVs into any
<SYMBOL>/ folder, its tabs light up automatically.

  * Pure Python standard library — NO Flask / pip install needed.
  * Fundamentals reuse fund_loader.load_stock(); fully OFFLINE.

RUN
    python stock_server.py              # http://localhost:8000
    python stock_server.py 8080         # choose a port
===============================================================================
"""

from __future__ import annotations

import json
import socket
import ssl
import sys
import threading
import time
import urllib.request
import webbrowser
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import urlparse, parse_qs, quote

import pandas as pd

import fund_loader

ROOT = Path(__file__).resolve().parent
STATEMENT_DIRS = ["pnl", "quarterly", "ratios", "balance_sheet", "cash_flow"]
MAX_EVENT_ROWS = 500          # cap rows sent per event feed (newest first)
FILE_CACHE = ROOT / "processed" / "nse_files"   # downloaded PDFs/XMLs from NSE
FILE_CACHE.mkdir(parents=True, exist_ok=True)


# ---------------------------------------------------------------------------
# Symbol discovery
# ---------------------------------------------------------------------------
def discover_symbols() -> list[str]:
    syms: set[str] = set()
    for st in STATEMENT_DIRS:
        d = ROOT / st
        if d.exists():
            syms.update(p.stem for p in d.glob("*.csv"))
    # Drop purely-numeric BSE scrip codes (e.g. 500142) — those are old/delisted
    # entries with stale, incomplete data. Keep proper alphabetic NSE symbols.
    return sorted(s for s in syms if not s.isdigit())


# ---------------------------------------------------------------------------
# Fundamentals payload
# ---------------------------------------------------------------------------
def _col(df, name):
    if df is None:
        return None
    for c in df.columns:
        if c.strip() == name:
            return c
    for c in df.columns:
        if name.lower() in c.lower():
            return c
    return None


def _num(v):
    try:
        if v is None or pd.isna(v):
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


def _latest(df, name):
    col = _col(df, name)
    if col is None:
        return None
    s = df[col].dropna()
    return _num(s.iloc[-1]) if not s.empty else None


def _series(df, name):
    col = _col(df, name)
    if df is None or col is None:
        return None
    return [_num(v) for v in df[col].tolist()]


def build_payload(sym: str) -> dict:
    data = fund_loader.load_stock(sym)
    pnl = data.get("pnl")
    q = data.get("quarterly")
    rat = data.get("ratios")
    bs = data.get("balance_sheet")

    eq, res, bor = (_latest(bs, "Equity Capital"),
                    _latest(bs, "Reserves"),
                    _latest(bs, "Borrowings"))
    de = None
    if None not in (eq, res, bor) and (eq + res):
        de = round(bor / (eq + res), 2)

    annual = {"labels": [], "sales": [], "net_profit": []}
    if pnl is not None and not pnl.empty:
        annual["labels"] = [d.strftime("%Y") for d in pnl.index]
        annual["sales"] = _series(pnl, "Sales") or []
        annual["net_profit"] = _series(pnl, "Net Profit") or []

    quarterly = {"labels": [], "net_profit": [], "yoy": []}
    if q is not None and not q.empty:
        sub = q.tail(12)
        quarterly["labels"] = [d.strftime("%b-%y") for d in sub.index]
        npc = _col(sub, "Net Profit")
        yoyc = _col(sub, "YOY Profit Growth %")
        quarterly["net_profit"] = [_num(v) for v in sub[npc].tolist()] if npc else []
        quarterly["yoy"] = [_num(v) for v in sub[yoyc].tolist()] if yoyc else []

    return {
        "symbol": sym,
        "have": [k for k in STATEMENT_DIRS if k in data],
        "cards": {
            "sales": _latest(pnl, "Sales"),
            "sales_g": _latest(pnl, "Sales Growth %"),
            "net_profit": _latest(pnl, "Net Profit"),
            "net_profit_g": _latest(pnl, "Profit Growth %"),
            "opm": _latest(pnl, "OPM %"),
            "roce": _latest(rat, "ROCE %"),
            "de": de,
        },
        "annual": annual,
        "quarterly": quarterly,
    }


# ---------------------------------------------------------------------------
# Event feeds payload  (announcements / board / corp-actions / results)
# ---------------------------------------------------------------------------
# feed key -> (nice label, date column for sorting, [(col, header, type)])
EVENT_FEEDS = {
    "announcements": {
        "label": "Announcements",
        "date": "sort_date",
        "columns": [
            ("sort_date", "Date", "date"),
            ("desc", "Category", "text"),
            ("attchmntText", "Details", "text"),
            ("attchmntFile", "File", "link"),
        ],
    },
    "board_meetings": {
        "label": "Board Meetings",
        "date": "bm_date",
        "columns": [
            ("bm_date", "Meeting date", "date"),
            ("bm_purpose", "Purpose", "text"),
            ("bm_desc", "Description", "text"),
            ("attachment", "File", "link"),
        ],
    },
    "corporate_actions": {
        "label": "Corporate Actions",
        "date": "exDate",
        "columns": [
            ("exDate", "Ex-date", "date"),
            ("recDate", "Record date", "date"),
            ("subject", "Subject", "text"),
            ("faceVal", "Face value", "text"),
        ],
    },
    "financial_results": {
        "label": "Financial Results",
        "date": "broadCastDate",
        "columns": [
            ("broadCastDate", "Broadcast", "date"),
            ("relatingTo", "Period", "text"),
            ("consolidated", "Basis", "text"),
            ("audited", "Audited", "text"),
            ("fromDate", "From", "date"),
            ("toDate", "To", "date"),
            ("xbrl", "XBRL", "link"),
        ],
    },
}


def _fmt_date(v: str) -> str:
    if not v:
        return ""
    ts = pd.to_datetime(v, errors="coerce", dayfirst=True, format="mixed")
    return ts.strftime("%d-%b-%Y") if pd.notna(ts) else str(v)[:20]


def build_events(sym: str) -> dict:
    """Read the four per-stock event CSVs from <SYMBOL>/ if present."""
    folder = ROOT / sym
    out = {"symbol": sym, "available": folder.exists(), "feeds": {}}

    for key, spec in EVENT_FEEDS.items():
        cols = spec["columns"]
        entry = {
            "label": spec["label"],
            "columns": [{"key": k, "label": lbl, "type": t} for k, lbl, t in cols],
            "rows": [], "total": 0, "shown": 0,
        }
        path = folder / f"{key}.csv"
        if path.exists():
            out["available"] = True
            try:
                df = pd.read_csv(path, dtype=str).fillna("")
            except Exception as e:                       # noqa: BLE001
                entry["error"] = f"{type(e).__name__}: {e}"
                out["feeds"][key] = entry
                continue
            dc = spec["date"]
            if dc in df.columns:
                df["_sort"] = pd.to_datetime(df[dc], errors="coerce", format="mixed")
                df = df.sort_values("_sort", ascending=False, na_position="last")
            entry["total"] = int(len(df))
            df = df.head(MAX_EVENT_ROWS)
            entry["shown"] = int(len(df))
            rows = []
            for _, r in df.iterrows():
                row = {}
                for k, _lbl, t in cols:
                    val = r[k] if k in df.columns else ""
                    row[k] = _fmt_date(val) if t == "date" else val
                rows.append(row)
            entry["rows"] = rows
        out["feeds"][key] = entry
    return out


def fetch_nse_file(url: str) -> tuple[bytes, str, str]:
    """Download an NSE attachment (PDF/XML) through the primed session, cached.

    NSE blocks/hangs on direct browser hits to these files, so the dashboard
    proxies them: we fetch with proper cookies+headers and stream the bytes back.
    """
    import mimetypes
    name = (url.rsplit("/", 1)[-1] or "file").split("?")[0]
    ctype = mimetypes.guess_type(name)[0] or "application/octet-stream"
    cache = FILE_CACHE / name

    if cache.exists() and cache.stat().st_size > 0:
        return cache.read_bytes(), ctype, name

    last = None
    for delay in (0, 2, 5):
        if delay:
            time.sleep(delay)
        with _NSE_LOCK:
            if _NSE_SESSION["op"] is None:
                _NSE_SESSION["op"] = _nse_prime("NIFTY")
            op = _NSE_SESSION["op"]
        try:
            with op.open(urllib.request.Request(url, headers=_NSE_HEADERS),
                         timeout=45) as r:
                data = r.read()
            if data:
                cache.write_bytes(data)
                return data, ctype, name
            last = RuntimeError("empty file from NSE")
        except Exception as e:                           # noqa: BLE001
            last = e
        with _NSE_LOCK:                                  # refresh session, retry
            _NSE_SESSION["op"] = None
    raise last or RuntimeError("could not download the file")


def xml_to_html(data: bytes, name: str, url: str) -> str:
    """Turn an NSE XBRL attachment (e.g. a board-meeting prior intimation) into a
    readable page instead of dumping raw XML at the user."""
    import html as _h
    import re
    import xml.etree.ElementTree as ET
    from urllib.parse import quote
    import xbrl_parser as XP

    root = ET.fromstring(data)
    skip = {"schemaRef", "context", "entity", "identifier", "period", "startDate",
            "endDate", "instant", "unit", "measure", "segment", "scenario",
            "explicitMember", "xbrl"}
    rows, seen = [], set()
    for el in root.iter():
        tag = XP.local(el.tag)
        if tag in skip or el.get("contextRef") is None:
            continue
        txt = (el.text or "").strip()
        if not txt or tag in seen:
            continue
        seen.add(tag)
        label = re.sub(r"(?<=[a-z0-9])(?=[A-Z])|(?<=[A-Z])(?=[A-Z][a-z])", " ", tag)
        rows.append((label, {"true": "Yes", "false": "No"}.get(txt.lower(), txt)))
    if not rows:
        raise RuntimeError("no readable fields in this XML")

    f = dict(rows)
    company = f.get("Name Of The Company", "")
    symbol = f.get("NSE Symbol", "")
    kind = f.get("Type Of Meeting") or f.get("Type Of Intimation") or "NSE Filing"
    body = "".join(f"<tr><td class='k'>{_h.escape(k)}</td>"
                   f"<td class='v'>{_h.escape(v)}</td></tr>" for k, v in rows)
    raw_link = f"/api/file?url={quote(url, safe='')}&amp;raw=1"
    raw_xml = _h.escape(data.decode("utf-8", "replace"))
    return f"""<!doctype html><html><head><meta charset="utf-8">
<title>{_h.escape(kind)} — {_h.escape(symbol or name)}</title><style>
body{{background:#0f131b;color:#e6edf3;margin:0;padding:34px;
 font-family:-apple-system,"Segoe UI",system-ui,sans-serif}}
.wrap{{max-width:880px;margin:0 auto}}
h1{{margin:0 0 4px;font-size:1.55rem;font-weight:700}}
.sub{{color:#8b98ad;font-family:ui-monospace,Consolas,monospace;font-size:.85rem;
 margin-bottom:22px}}
table{{width:100%;border-collapse:collapse;background:#171d2b;border:1px solid #2a3346;
 border-radius:10px;overflow:hidden}}
td{{padding:11px 16px;border-bottom:1px solid #2a3346;vertical-align:top}}
tr:last-child td{{border-bottom:0}} tr:hover{{background:#1b2333}}
.k{{color:#8b98ad;width:44%;font-size:.9rem}}
.v{{color:#e6edf3;font-weight:600}}
a{{color:#58a6ff;font-family:ui-monospace,Consolas,monospace;font-size:.82rem;
 text-decoration:none}}
a:hover{{text-decoration:underline}}
.btn{{display:inline-block;background:#58a6ff;color:#0b1220;font-weight:700;
 padding:8px 15px;border-radius:8px;margin-right:10px}}
.btn:hover{{background:#7cb8ff;text-decoration:none}}
.ghost{{display:inline-block;background:#171d2b;border:1px solid #2a3346;color:#e6edf3;
 padding:8px 15px;border-radius:8px}}
.ghost:hover{{background:#243044;text-decoration:none}}
.foot{{margin-top:18px}}
details{{margin-top:20px;background:#171d2b;border:1px solid #2a3346;border-radius:10px}}
summary{{cursor:pointer;padding:11px 16px;color:#8b98ad;
 font-family:ui-monospace,Consolas,monospace;font-size:.85rem}}
summary:hover{{color:#e6edf3}}
pre{{margin:0;padding:16px;overflow-x:auto;max-height:60vh;overflow-y:auto;
 border-top:1px solid #2a3346;font-family:ui-monospace,Consolas,monospace;
 font-size:12px;line-height:1.5;color:#8b98ad;white-space:pre}}
</style></head><body><div class="wrap">
<h1>{_h.escape(kind)}</h1>
<div class="sub">{_h.escape(company)}{(' · ' + _h.escape(symbol)) if symbol else ''}</div>
<table>{body}</table>
<div class="foot">
  <a class="btn" href="{raw_link}&amp;dl=1">⬇ Download XML file</a>
  <a class="ghost" href="{raw_link}" target="_blank">Open raw XML in new tab ↗</a>
</div>
<details><summary>▸ Show raw XML ({len(data):,} bytes)</summary><pre>{raw_xml}</pre></details>
</div></body></html>"""


def _xbrl_facts(xml_text: str) -> list[tuple]:
    """Every non-dimensional numeric fact: (concept, value, period_from, period_to)."""
    import xml.etree.ElementTree as ET
    import xbrl_parser as XP
    root = ET.fromstring(xml_text)
    ctx = XP.parse_contexts(root)
    out = []
    for el in root.iter():
        cid = el.get("contextRef")
        if not cid or cid not in ctx:
            continue
        start, end, dim = ctx[cid]
        if dim:                                   # skip per-segment contexts
            continue
        txt = (el.text or "").strip()
        if not txt:
            continue
        try:
            val = float(txt.replace(",", ""))
        except ValueError:
            continue
        out.append((XP.local(el.tag), val, start, end))
    return out


def build_xbrl_excel(sym: str, url: str) -> tuple[bytes, str]:
    """Download an NSE XBRL filing and turn it into a formatted .xlsx workbook."""
    import io
    import xbrl_parser as XP
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    # metadata for this filing (period dates, basis, etc.) from the feed
    meta = {}
    path = ROOT / sym / "financial_results.csv"
    if path.exists():
        df = pd.read_csv(path, dtype=str).fillna("")
        if "xbrl" in df.columns:
            m = df[df["xbrl"] == url]
            if not m.empty:
                meta = m.iloc[0].to_dict()

    text = XP.download(url)
    if not text:
        raise RuntimeError("could not download the XBRL file from NSE")

    p_from, p_to = meta.get("fromDate", ""), meta.get("toDate", "")
    key = XP.extract(text, p_from, p_to) if (p_from and p_to) else {}
    if "_error" in key:
        key = {}
    facts = _xbrl_facts(text)
    if not key and not facts:
        raise RuntimeError("no figures found (older Non-Ind-AS schema)")

    wb = Workbook()
    hdr = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F4E79")
    bold = Font(bold=True)

    # ---- sheet 1: key figures --------------------------------------------
    ws = wb.active
    ws.title = "Key Figures"
    r = 1
    for k, v in [("Company", meta.get("companyName", sym)), ("Symbol", sym),
                 ("Period", meta.get("relatingTo", "")),
                 ("From", p_from), ("To", p_to),
                 ("Basis", meta.get("consolidated", "")),
                 ("Audited", meta.get("audited", "")),
                 ("Broadcast", meta.get("broadCastDate", "")),
                 ("Source", url)]:
        ws.cell(r, 1, k).font = bold
        ws.cell(r, 2, v)
        r += 1
    r += 1
    ws.cell(r, 1, "KEY FIGURES").font = bold
    r += 1
    for i, h in enumerate(["Metric", "Value"], 1):
        c = ws.cell(r, i, h)
        c.font = hdr
        c.fill = fill
    r += 1
    for k, label, div in [("revenue", "Revenue from Operations (₹ cr)", 1e7),
                          ("other_income", "Other Income (₹ cr)", 1e7),
                          ("total_income", "Total Income (₹ cr)", 1e7),
                          ("profit_before_tax", "Profit Before Tax (₹ cr)", 1e7),
                          ("net_profit", "Net Profit (₹ cr)", 1e7),
                          ("basic_eps", "Basic EPS (₹)", 1)]:
        ws.cell(r, 1, label)
        v = key.get(k)
        if v is None:
            ws.cell(r, 2, "—")
        else:
            c = ws.cell(r, 2, round(v / div, 2))
            c.number_format = "#,##0.00"
        r += 1
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 52

    # ---- sheet 2: every reported figure -----------------------------------
    ws2 = wb.create_sheet("All Figures")
    for i, h in enumerate(["Concept", "Value (as reported)", "Value (₹ crore)",
                           "Period from", "Period to"], 1):
        c = ws2.cell(1, i, h)
        c.font = hdr
        c.fill = fill
    rr = 2
    for name, val, s, e in facts:
        ws2.cell(rr, 1, name)
        c = ws2.cell(rr, 2, val)
        c.number_format = "#,##0.00"
        c = ws2.cell(rr, 3, round(val / 1e7, 4))
        c.number_format = "#,##0.00"
        ws2.cell(rr, 4, s)
        ws2.cell(rr, 5, e)
        rr += 1
    for col, w in zip("ABCDE", [56, 22, 18, 14, 14]):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    per = (meta.get("relatingTo") or "Result").replace(" ", "")
    basis = (meta.get("consolidated") or "").replace("-", "")
    end = (p_to or "").replace(" ", "")[:11]
    fname = f"{sym}_{per}_{basis}_{end}.xlsx".replace("/", "-").replace("\\", "-")
    return buf.getvalue(), fname


# folders that hold the offline financial statements (Screener-style CSVs)
STATEMENT_SHEETS = [
    ("pnl",           "P&L"),
    ("balance_sheet", "Balance Sheet"),
    ("cash_flow",     "Cash Flow"),
    ("ratios",        "Ratios"),
    ("quarterly",     "Quarterly"),
]


def build_statements_excel(sym: str) -> tuple[bytes, str]:
    """Build a .xlsx of the financial statements from the LOCAL folders
    (pnl / balance_sheet / cash_flow / ratios / quarterly), NOT from any online
    source. One sheet per statement, mirroring the folder CSV exactly."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    hdr = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F4E79")
    metric_font = Font(bold=True)

    wb = Workbook()
    wb.remove(wb.active)                       # start with no sheets; add what exists
    found = []
    for folder, title in STATEMENT_SHEETS:
        path = ROOT / folder / f"{sym}.csv"
        if not path.exists():
            continue
        df = pd.read_csv(path, dtype=str).fillna("")
        ws = wb.create_sheet(title[:31])
        # header row
        for j, col in enumerate(df.columns, 1):
            c = ws.cell(1, j, col)
            c.font = hdr
            c.fill = fill
        # data rows — numbers become real numbers so Excel can chart/sum them
        for i, (_, row) in enumerate(df.iterrows(), 2):
            for j, col in enumerate(df.columns, 1):
                v = row[col]
                if j == 1:                     # the "Metric" label column
                    c = ws.cell(i, j, v)
                    c.font = metric_font
                    continue
                num = _num(v)                  # '1,234' / '-5%' -> float, else NaN
                if v not in ("", "-") and num == num:
                    c = ws.cell(i, j, num)
                    c.number_format = ('0.00"%"' if "%" in str(v) else "#,##0.00")
                else:
                    ws.cell(i, j, v)
        ws.column_dimensions["A"].width = 34
        for col_cells in ws.iter_cols(min_col=2, max_col=ws.max_column):
            ws.column_dimensions[col_cells[0].column_letter].width = 12
        ws.freeze_panes = "B2"
        found.append(folder)

    if not found:
        raise RuntimeError(
            f"no local statement files for '{sym}' — expected e.g. pnl/{sym}.csv. "
            f"(the offline folders key files by NSE symbol; this stock may only "
            f"exist under a BSE code.)")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), f"{sym}_financials.xlsx"


def downloaded_symbols() -> list[str]:
    """Symbols whose <sym>/ folder already holds at least one event CSV."""
    files = [f"{k}.csv" for k in EVENT_FEEDS]
    ready = []
    for s in SYMBOLS:
        folder = ROOT / s
        if folder.exists() and any((folder / f).exists() for f in files):
            ready.append(s)
    return ready


def _avg_path(px, dates, N, M) -> list:
    """Average % price path around the event (day 0 = event day = 0%)."""
    idx = px.index
    rows = []
    for ev in dates:
        pos = idx.searchsorted(pd.Timestamp(ev))
        if pos - N < 0 or pos + M >= len(idx):
            continue
        base = float(px.iloc[pos])
        if base <= 0:
            continue
        w = px.iloc[pos - N:pos + M + 1].values
        rows.append([(float(v) / base - 1) * 100 for v in w])
    if not rows:
        return []
    return [round(sum(col) / len(col), 3) for col in zip(*rows)]


def _last_event_window(px, dates, before=6, after=3) -> dict | None:
    """Actual prices around the MOST RECENT event (day 0 = event day)."""
    idx = px.index
    best = None
    for d in sorted(dates):
        pos = idx.searchsorted(pd.Timestamp(d))
        if pos - before < 0 or pos + after >= len(idx):
            continue
        w = px.iloc[pos - before:pos + after + 1]
        best = {"date": str(idx[pos].date()),
                "days": list(range(-before, after + 1)),
                "prices": [round(float(v), 2) for v in w.values]}
    return best


def _all_event_windows(px, dates, before=6, after=3) -> list:
    """Every PAST event (newest first) with the same before→event→after summary
    the 'recent event report' shows — so the whole history can be browsed, not
    just the latest one. Day 0 = event day."""
    idx = px.index
    out = []
    for d in sorted(dates):
        pos = idx.searchsorted(pd.Timestamp(d))
        if pos - before < 0 or pos + after >= len(idx):
            continue
        w = px.iloc[pos - before:pos + after + 1]
        p0 = float(w.iloc[0])                      # `before` days before (buy)
        pe = float(px.iloc[pos])                   # event day
        pa = float(w.iloc[-1])                     # `after` days after (sell)
        if p0 <= 0 or pe <= 0:
            continue
        out.append({
            "date": str(idx[pos].date()),
            "buy": round(p0, 2), "event": round(pe, 2), "sell": round(pa, 2),
            "before_pct": round((pe / p0 - 1) * 100, 2),
            "after_pct": round((pa / pe - 1) * 100, 2),
            "total_pct": round((pa / p0 - 1) * 100, 2),
            "days": list(range(-before, after + 1)),
            "prices": [round(float(v), 2) for v in w.values],
        })
    out.reverse()                                  # newest first
    return out


def build_behaviour(sym: str) -> dict:
    """Event behaviour grid for one stock: for each event_type, the avg return %
    and win-rate for every (buy T-N days before, sell T+M days after) pair, plus
    the single best rule. Reuses event_behaviour.py + cached Yahoo prices."""
    import event_behaviour as EB

    events = EB.load_events_from_feeds([sym])
    if events.empty:
        return {"symbol": sym, "available": False, "reason": "no event feeds"}
    px = EB.fetch_prices_yahoo(sym)
    if px is None or len(px) < (EB.N_BEFORE + EB.N_AFTER + 5):
        return {"symbol": sym, "available": False, "reason": "no price history"}

    rows = []
    for _, e in events.iterrows():
        rows += EB.analyse_event(px, e["event_date"], sym, e["event_type"])
    if not rows:
        return {"symbol": sym, "available": False, "reason": "events outside price range"}

    detail = pd.DataFrame(rows)
    agg = EB.aggregate(detail)
    NB, NA = EB.N_BEFORE, EB.N_AFTER

    types = {}
    for et, sub in agg.groupby("event_type"):
        gr = [[None] * NA for _ in range(NB)]
        gw = [[None] * NA for _ in range(NB)]
        for _, r in sub.iterrows():
            db, da = int(r["days_before_entry"]) - 1, int(r["post_event_day"]) - 1
            gr[db][da] = round(float(r["avg_return_pct"]), 2)
            gw[db][da] = round(float(r["win_rate_pct"]), 0)
        n_max = int(sub["n_events"].max())
        keep = sub[sub["n_events"] >= max(3, min(5, n_max))]
        if keep.empty:                      # very few events -> don't filter them out
            keep = sub
        # prefer the highest win-rate among profitable rules (most intuitive),
        # fall back to best average if nothing is profitable.
        pos = keep[keep["avg_return_pct"] > 0]
        pool = pos if not pos.empty else keep
        if pool.empty:                      # nothing to pick (shouldn't happen) -> skip type
            continue
        b = pool.sort_values(["win_rate_pct", "avg_return_pct"], ascending=False).iloc[0]

        # outcome distribution at the chosen rule
        cell = detail[(detail["event_type"] == et) &
                      (detail["days_before_entry"] == int(b["days_before_entry"])) &
                      (detail["post_event_day"] == int(b["post_event_day"]))]["return_pct"]
        dates = events[events["event_type"] == et]["event_date"]

        # ── HONEST out-of-sample check: choose the best rule on the FIRST half of
        # events, then score that rule's win-rate on the (unseen) SECOND half.
        # Purely additive — the in-sample win_rate_pct above is untouched. ──────
        oos_win = oos_n = None
        det_et = detail[detail["event_type"] == et]
        edates = sorted(det_et["event_date"].unique())
        if len(edates) >= 8:
            mid = edates[len(edates) // 2]
            tr = det_et[det_et["event_date"] <= mid]
            teo = det_et[det_et["event_date"] > mid]
            g = tr.groupby(["days_before_entry", "post_event_day"])
            st = g["win_loss"].apply(lambda s: (s == "WIN").mean()).reset_index(name="w")
            st["n"] = g.size().values
            st = st[st["n"] >= max(2, min(3, int(st["n"].max())))]
            if len(st):
                bb = st.sort_values("w", ascending=False).iloc[0]
                c = teo[(teo["days_before_entry"] == int(bb["days_before_entry"])) &
                        (teo["post_event_day"] == int(bb["post_event_day"]))]["win_loss"]
                if len(c):
                    oos_win = round((c == "WIN").mean() * 100)
                    oos_n = int(len(c))

        types[str(et)] = {
            "n_events": n_max,
            "grid_return": gr, "grid_win": gw,
            "path": {"days": list(range(-NB, NA + 1)), "avg_pct": _avg_path(px, dates, NB, NA)},
            "last": _last_event_window(px, dates, 6, 3),
            "history": _all_event_windows(px, dates, 6, 3),   # every past event, newest first
            "best": {"days_before": int(b["days_before_entry"]),
                     "days_after": int(b["post_event_day"]),
                     "avg_return_pct": round(float(b["avg_return_pct"]), 2),
                     "win_rate_pct": round(float(b["win_rate_pct"]), 0),
                     "oos_win": oos_win, "oos_n": oos_n,
                     "median_pct": round(float(cell.median()), 2) if len(cell) else None,
                     "best_case": round(float(cell.max()), 2) if len(cell) else None,
                     "worst_case": round(float(cell.min()), 2) if len(cell) else None,
                     "n": int(b["n_events"])},
        }
    return {"symbol": sym, "available": True, "n_before": NB, "n_after": NA, "types": types}


# ---------------------------------------------------------------------------
# Disk cache — persist the expensive rankings/screener payloads so a restart (or
# a first tab-load) serves instantly instead of recomputing for ~30-40 s. The
# cache key folds in a data fingerprint (the latest mtime of the feed/price
# files), so it auto-refreshes whenever the underlying data actually changes.
# ---------------------------------------------------------------------------
_CACHE_DIR = ROOT / "processed" / "cache"


def _disk_cached(name: str, universe, paths, compute):
    import hashlib
    key = hashlib.md5("|".join(universe).encode()).hexdigest()[:10]
    fp = 0
    for p in paths:
        try:
            fp = max(fp, int(p.stat().st_mtime))
        except OSError:
            pass
    path = _CACHE_DIR / f"{name}_{key}_{fp}.json"
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception:                          # noqa: BLE001
            pass
    result = compute()
    try:
        _CACHE_DIR.mkdir(parents=True, exist_ok=True)
        for old in _CACHE_DIR.glob(f"{name}_{key}_*.json"):   # drop stale versions
            try:
                old.unlink()
            except OSError:
                pass
        path.write_text(json.dumps(result, ensure_ascii=False), encoding="utf-8")
    except Exception:                              # noqa: BLE001
        pass
    return result


# ---------------------------------------------------------------------------
# Rankings — every stock ordered by win-rate, per event type (cached)
# ---------------------------------------------------------------------------
RANK_EVENTS = ["RESULTS", "BOARD_MEETING", "CORPORATE_ACTION", "ANNOUNCEMENT"]
_RANK_CACHE: dict = {}


def build_rankings() -> dict:
    """For each event type, all stocks (with feeds) ranked by the best rule's
    win-rate, highest first. Reuses build_behaviour; cached in memory + on disk."""
    universe = [s for s in SYMBOLS if s in set(downloaded_symbols())]
    key = tuple(universe)
    if key in _RANK_CACHE:
        return _RANK_CACHE[key]

    def _compute():
        import event_behaviour as EB
        out = {et: [] for et in RANK_EVENTS}
        for sym in universe:
            try:
                d = build_behaviour(sym)
            except Exception:                      # noqa: BLE001
                continue
            if not d.get("available"):
                continue
            for et, t in d["types"].items():
                b = t.get("best")
                if not b:
                    continue
                out.setdefault(et, []).append({
                    "symbol": sym,
                    "win": b["win_rate_pct"], "avg": b["avg_return_pct"],
                    "oos": b.get("oos_win"),
                    "db": b["days_before"], "da": b["days_after"],
                    "n": b["n"], "events": t.get("n_events", b["n"]),
                })
        for et in out:
            out[et].sort(key=lambda r: (r["win"], r["avg"]), reverse=True)
        return {"events": RANK_EVENTS, "ranks": out, "n_stocks": len(universe)}

    import event_behaviour as EB
    paths = ([ROOT / s / f"{k}.csv" for s in universe for k in EVENT_FEEDS]
             + [EB.PRICE_CACHE / f"{s}.csv" for s in universe])
    result = _disk_cached("rankings", universe, paths, _compute)
    _RANK_CACHE[key] = result
    return result


# ---------------------------------------------------------------------------
# Screener — latest fundamentals for every stock (cached)
# ---------------------------------------------------------------------------
_SCREEN_CACHE: dict = {}


def build_screener() -> dict:
    """Latest annual fundamentals per stock — ROCE, growth, margins, debt, P/E —
    for the Screener and Compare tabs. Reuses fund_loader + cached prices."""
    import fund_loader
    import event_behaviour as EB

    universe = [s for s in SYMBOLS if (ROOT / "pnl" / f"{s}.csv").exists()]
    key = tuple(universe)
    if key in _SCREEN_CACHE:
        return _SCREEN_CACHE[key]

    def f(v):
        try:
            v = float(v)
            return round(v, 2) if v == v else None
        except (TypeError, ValueError):
            return None

    def _compute():
        out = []
        for sym in universe:
            try:
                d = fund_loader.load_stock(sym)
            except Exception:                      # noqa: BLE001
                continue
            pnl, bs, rat = d.get("pnl"), d.get("balance_sheet"), d.get("ratios")
            if pnl is None or pnl.empty:
                continue
            L = pnl.iloc[-1]
            eps = f(L.get("EPS in Rs"))
            px = EB.fetch_prices_yahoo(sym)
            price = f(px.iloc[-1]) if (px is not None and len(px)) else None
            pe = round(price / eps, 1) if (price and eps and eps > 0) else None
            roce = f(rat.iloc[-1].get("ROCE %")) if (rat is not None and not rat.empty) else None
            de = None
            if bs is not None and not bs.empty:
                B = bs.iloc[-1]
                bor, eq, res = B.get("Borrowings"), B.get("Equity Capital"), B.get("Reserves")
                try:
                    base = (float(eq) if eq == eq else 0) + (float(res) if res == res else 0)
                    if bor == bor and base:
                        de = round(float(bor) / base, 2)
                except (TypeError, ValueError):
                    pass
            out.append({
                "symbol": sym, "price": price, "pe": pe, "roce": roce,
                "opm": f(L.get("OPM %")), "sales_g": f(L.get("Sales Growth %")),
                "profit_g": f(L.get("Profit Growth %")), "de": de,
                "sales": f(L.get("Sales")), "npat": f(L.get("Net Profit")), "eps": eps,
            })
        return {"stocks": out, "n": len(out)}

    paths = ([ROOT / d / f"{s}.csv" for s in universe for d in ("pnl", "ratios", "balance_sheet")]
             + [EB.PRICE_CACHE / f"{s}.csv" for s in universe])
    result = _disk_cached("screener", universe, paths, _compute)
    _SCREEN_CACHE[key] = result
    return result


# ---------------------------------------------------------------------------
# Upcoming events — an actionable watchlist of events that are still ahead,
# each carrying the stock's historical best rule + win-rate (in & out of sample).
# ---------------------------------------------------------------------------
_UP_FEEDS = {"RESULTS": ("financial_results", "broadCastDate"),
             "BOARD_MEETING": ("board_meetings", "bm_date"),
             "CORPORATE_ACTION": ("corporate_actions", "exDate")}


def build_upcoming(days: int = 120) -> dict:
    import datetime as _dt
    ranks = build_rankings()                       # cached; gives best rule + oos per stock/event
    bestmap = {(r["symbol"], et): r for et, lst in ranks["ranks"].items() for r in lst}
    today = pd.Timestamp(_dt.date.today())
    horizon = today + pd.Timedelta(days=days)

    def row(sym, et, d, status):
        r = bestmap.get((sym, et), {})
        return {"symbol": sym, "event": et, "date": str(pd.Timestamp(d).date()),
                "in_days": int((pd.Timestamp(d).normalize() - today).days), "status": status,
                "db": r.get("db"), "da": r.get("da"), "win": r.get("win"),
                "oos": r.get("oos"), "avg": r.get("avg"), "events": r.get("events")}

    rows = []
    for sym in SYMBOLS:
        for et, (f, dc) in _UP_FEEDS.items():
            p = ROOT / sym / f"{f}.csv"
            if not p.exists():
                continue
            df = pd.read_csv(p, dtype=str).fillna("")
            if dc not in df.columns:
                continue
            dts = pd.to_datetime(df[dc], errors="coerce", format="mixed").dropna().dt.normalize()
            if dts.empty:
                continue
            for x in sorted(dts[(dts > today) & (dts <= horizon)].unique()):   # confirmed
                rows.append(row(sym, et, x, "confirmed"))
            # estimate the NEXT results / corporate action from the stock's cadence
            if et in ("RESULTS", "CORPORATE_ACTION"):
                past = dts[dts <= today].sort_values()
                if len(past) >= 4:
                    med = past.diff().dropna().dt.days.median()
                    if med and med == med:
                        nxt = past.iloc[-1] + pd.Timedelta(days=float(med))
                        if today < nxt <= horizon and not ((dts > today) & (dts <= horizon)).any():
                            rows.append(row(sym, et, nxt, "estimated"))
    rows.sort(key=lambda r: (r["date"], r["symbol"]))
    return {"today": str(today.date()), "horizon_days": days, "n": len(rows), "rows": rows}


# ---------------------------------------------------------------------------
# NSE downloader — fetch the four event feeds for ANY symbol on demand
# ---------------------------------------------------------------------------
_NSE_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/122.0 Safari/537.36",
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://www.nseindia.com/companies-listing/corporate-filings-announcements",
}
_NSE_CTX = ssl.create_default_context()
_NSE_CTX.check_hostname = False
_NSE_CTX.verify_mode = ssl.CERT_NONE

NSE_API = {
    "announcements":     "https://www.nseindia.com/api/corporate-announcements?index=equities&symbol={s}",
    "board_meetings":    "https://www.nseindia.com/api/corporate-board-meetings?index=equities&symbol={s}",
    "corporate_actions": "https://www.nseindia.com/api/corporates-corporateActions?index=equities&symbol={s}",
}
NSE_RESULTS = "https://www.nseindia.com/api/corporates-financial-results?index=equities&symbol={s}&period={p}"


# ONE long-lived, cookie-primed NSE session, reused across every download so we
# don't re-hit NSE's homepage (and trip its bot-wall) on every request.
_NSE_SESSION = {"op": None}
_NSE_LOCK = threading.Lock()


def _nse_prime(sym: str):
    """Build a fresh cookie-primed opener (homepage + the stock's quote page)."""
    op = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(),
                                     urllib.request.HTTPSHandler(context=_NSE_CTX))
    for u in ("https://www.nseindia.com/",
              f"https://www.nseindia.com/get-quotes/equity?symbol={sym}"):
        try:
            op.open(urllib.request.Request(u, headers=_NSE_HEADERS), timeout=25).read()
        except Exception:                                # noqa: BLE001 - keep warming
            pass
        time.sleep(0.6)
    return op


def _nse_rows(sym: str, url: str) -> list:
    """GET one NSE json api, reusing the shared session; re-prime + back off on
    rejection (403 / empty). Raises the last error only if every attempt fails."""
    last = None
    for delay in (0, 2, 5):                              # exponential-ish backoff
        if delay:
            time.sleep(delay)
        with _NSE_LOCK:
            if _NSE_SESSION["op"] is None:
                _NSE_SESSION["op"] = _nse_prime(sym)
            op = _NSE_SESSION["op"]
        try:
            raw = op.open(urllib.request.Request(url, headers=_NSE_HEADERS),
                          timeout=25).read().decode("utf-8", "replace").strip()
            if raw and raw[0] in "[{":
                d = json.loads(raw)
                return d if isinstance(d, list) else d.get("data", [])
            last = ValueError("empty response from NSE")
        except Exception as e:                           # noqa: BLE001
            last = e
        with _NSE_LOCK:                                  # session went bad -> refresh
            _NSE_SESSION["op"] = None
    if last:
        raise last
    return []


def fetch_nse_feeds(sym: str) -> dict:
    """Download the four event feeds for `sym` -> <sym>/*.csv.

    Every feed is isolated: one feed failing (403/empty) never kills the others,
    and whatever succeeds is saved. Returns per-feed counts + any errors.
    """
    folder = ROOT / sym
    folder.mkdir(exist_ok=True)
    counts, errors = {}, {}

    # Symbols like "M&M" and "BAJAJ-AUTO" must be percent-encoded, or the bare "&"
    # terminates the query string and NSE silently reads the symbol as just "M".
    esc = quote(sym, safe="")

    for key, tpl in NSE_API.items():
        try:
            r = _nse_rows(sym, tpl.format(s=esc))
            if r:
                pd.DataFrame(r).to_csv(folder / f"{key}.csv", index=False)
            counts[key] = len(r)
        except Exception as e:                           # noqa: BLE001
            errors[key] = f"{type(e).__name__}: {e}"
            counts[key] = 0
        time.sleep(0.5)

    fr = []
    for period in ("Quarterly", "Annual"):
        try:
            fr += _nse_rows(sym, NSE_RESULTS.format(s=esc, p=period))
        except Exception as e:                           # noqa: BLE001
            errors[f"results_{period.lower()}"] = f"{type(e).__name__}: {e}"
        time.sleep(0.5)
    if fr:
        df = pd.DataFrame(fr)
        try:
            df = df.drop_duplicates()
        except Exception:                                # noqa: BLE001
            pass
        df.to_csv(folder / "financial_results.csv", index=False)
    counts["financial_results"] = len(fr)

    return {"counts": counts, "errors": errors, "any": any(counts.values())}


# ---------------------------------------------------------------------------
# The single-page front end (inline HTML + CSS + JS, canvas charts, no CDN)
# ---------------------------------------------------------------------------
PAGE = r"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>NSE Stock Browser</title>
<style>
  :root{
    --bg:#0f131b; --panel:#171d2b; --grid:#2a3346; --ink:#e6edf3; --mute:#8b98ad;
    --pos:#3fb950; --neg:#f85149; --blue:#58a6ff; --amber:#d29922;
    --mono:ui-monospace,"Cascadia Code",Consolas,monospace;
    --sans:-apple-system,"Segoe UI",system-ui,sans-serif;
  }
  *{box-sizing:border-box}
  body{margin:0;background:var(--bg);color:var(--ink);font-family:var(--sans)}
  .bar{display:flex;align-items:center;gap:10px;flex-wrap:wrap;
       padding:14px 20px;border-bottom:1px solid var(--grid);position:sticky;top:0;
       background:var(--bg);z-index:5}
  .bar h1{font-size:1.15rem;margin:0;font-weight:650;margin-right:auto}
  button{background:var(--panel);color:var(--ink);border:1px solid var(--grid);
    border-radius:7px;padding:7px 12px;font-family:var(--sans);font-weight:600;
    font-size:.9rem;cursor:pointer}
  button:hover{background:#243044}
  button:active{transform:translateY(1px)}
  input{background:var(--panel);color:var(--ink);border:1px solid var(--grid);
    border-radius:7px;padding:7px 12px;font-family:var(--mono);font-size:1rem;
    width:200px;text-transform:uppercase}
  input::placeholder{text-transform:none;color:var(--mute)}
  input:focus{outline:2px solid var(--blue)}
  .searchwrap{position:relative;display:inline-block}
  .ac{position:absolute;top:calc(100% + 4px);left:0;width:260px;max-height:340px;
    overflow-y:auto;background:var(--panel);border:1px solid var(--grid);border-radius:8px;
    box-shadow:0 10px 30px rgba(0,0,0,.5);z-index:30;display:none}
  .acitem{padding:8px 12px;font-family:var(--mono);font-size:.9rem;cursor:pointer;
    color:var(--mute);border-bottom:1px solid #1e2636;white-space:nowrap;
    display:flex;justify-content:space-between;align-items:center;gap:12px}
  .acitem:last-child{border-bottom:0}
  .acitem b{color:var(--blue);font-weight:700}
  .acitem.sel,.acitem:hover{background:#243044;color:var(--ink)}
  .rdy{color:var(--pos);font-size:.68rem;font-weight:700;white-space:nowrap}
  .notrdy{color:var(--mute);font-size:.68rem;white-space:nowrap;opacity:.7}
  #evstatus{margin-left:10px;font-family:var(--mono);font-size:.75rem}
  .acempty{padding:10px 12px;color:var(--mute);font-family:var(--mono);font-size:.85rem}
  .acmore{padding:7px 12px;color:var(--mute);font-family:var(--mono);font-size:.75rem;
    border-top:1px solid var(--grid)}
  .counter{font-family:var(--mono);color:var(--mute);font-size:.9rem;min-width:110px;
    text-align:right}
  main{max-width:1040px;margin:0 auto;padding:18px 20px 40px}
  .sym{font-size:2rem;font-weight:700;letter-spacing:-.01em}
  .have{font-family:var(--mono);color:var(--mute);font-size:.8rem;margin-left:10px}

  /* tabs */
  .tabs{display:flex;gap:4px;flex-wrap:wrap;margin:14px 0 16px;
        border-bottom:1px solid var(--grid)}
  .tab{background:transparent;border:0;border-bottom:2px solid transparent;
       color:var(--mute);padding:9px 14px;font-weight:600;font-size:.9rem;
       cursor:pointer;border-radius:0}
  .tab:hover{background:transparent;color:var(--ink)}
  .tab.active{color:var(--blue);border-bottom-color:var(--blue)}

  .cards{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin:0 0 18px}
  @media(max-width:720px){.cards{grid-template-columns:repeat(2,1fr)}}
  .card{background:var(--panel);border:1px solid var(--grid);border-radius:10px;
    padding:14px 16px}
  .card .t{color:var(--mute);font-size:.8rem}
  .card .v{font-size:1.5rem;font-weight:700;margin:2px 0}
  .card .s{color:var(--mute);font-size:.75rem;font-family:var(--mono)}
  .charts{display:grid;grid-template-columns:1fr 1fr;gap:12px}
  @media(max-width:720px){.charts{grid-template-columns:1fr}}
  .chart{background:var(--panel);border:1px solid var(--grid);border-radius:10px;padding:12px}
  .chart .ct{font-weight:650;font-size:.95rem}
  .chart .cs{color:var(--mute);font-size:.78rem;float:right;font-family:var(--mono)}
  canvas{width:100%;height:300px;display:block;margin-top:6px}

  /* event tables */
  .tblwrap{border:1px solid var(--grid);border-radius:10px;overflow:auto;max-height:62vh;
    background:var(--panel)}
  table{border-collapse:collapse;width:100%;font-size:.86rem}
  thead th{position:sticky;top:0;background:#1c2434;color:var(--mute);
    font-family:var(--mono);font-size:.72rem;letter-spacing:.04em;text-transform:uppercase;
    text-align:left;padding:9px 12px;border-bottom:1px solid var(--grid);white-space:nowrap}
  td{padding:9px 12px;border-bottom:1px solid var(--grid);vertical-align:top}
  tbody tr:hover{background:#1b2333}
  td:nth-child(1){white-space:nowrap;font-family:var(--mono);color:var(--ink)}
  .cat{font-family:var(--mono);font-size:.78rem;color:var(--amber)}
  .enote{color:var(--mute);font-size:.85rem;padding:10px 4px;font-family:var(--mono)}
  /* behaviour heatmap */
  .bhead{font-size:1.05rem;margin:20px 0 4px;font-weight:650}
  .best{background:var(--panel);border:1px solid var(--grid);border-left:3px solid var(--pos);
    border-radius:0 8px 8px 0;padding:8px 12px;margin:4px 0 8px;font-size:.9rem}
  .best b{color:var(--ink)}
  table.hm{border-collapse:collapse;font-variant-numeric:tabular-nums}
  table.hm th{background:var(--panel);color:var(--mute);font-family:var(--mono);
    font-size:.72rem;padding:6px 8px;position:static;border:1px solid var(--grid);white-space:nowrap}
  table.hm td{padding:6px 9px;text-align:right;font-family:var(--mono);font-size:.8rem;
    color:var(--ink);border:1px solid var(--grid);min-width:46px}
  table.hm .bcell{outline:2px solid var(--blue);outline-offset:-2px;font-weight:800}
  .bwrap{overflow-x:auto;margin:4px 0 6px}
  /* simple behaviour cards */
  .bintro{color:var(--mute);font-size:.95rem;margin:6px 0 16px;line-height:1.5}
  .bintro b{color:var(--ink)}
  .bcards{display:grid;gap:14px;grid-template-columns:1fr 1fr}
  @media(max-width:720px){.bcards{grid-template-columns:1fr}}
  .bcard{background:var(--panel);border:1px solid var(--grid);border-radius:12px;padding:18px 20px}
  .bcard-top{display:flex;justify-content:space-between;align-items:baseline;margin-bottom:10px}
  .bet{font-weight:700;font-size:1.1rem}
  .bn{font-family:var(--mono);font-size:.72rem;color:var(--mute)}
  .brec{font-size:1.05rem;margin:2px 0 16px;line-height:1.5}
  .brec b{color:var(--blue)}
  .bstats{display:flex;gap:22px;align-items:flex-end;flex-wrap:wrap}
  .bstat{flex:1;min-width:130px}
  .blabel{font-size:.7rem;color:var(--mute);text-transform:uppercase;letter-spacing:.05em;
    margin-bottom:6px;font-family:var(--mono)}
  .bwinbar{position:relative;height:14px;background:#0e1420;border:1px solid var(--grid);
    border-radius:100px;overflow:hidden}
  .bwinbar>span{display:block;height:100%;border-radius:100px}
  .bwinbar::after{content:"";position:absolute;left:50%;top:0;bottom:0;width:1px;
    background:var(--ink);opacity:.35}
  .bwinval{font-family:var(--mono);font-weight:700;font-size:1.25rem;margin-top:6px}
  .bprofit{font-family:var(--mono);font-weight:700;font-size:1.6rem;line-height:1}
  .bverdict{margin-top:16px;padding:9px 13px;border-radius:8px;font-size:.9rem;font-weight:600}
  .v-strong{background:rgba(63,185,80,.14);color:var(--pos)}
  .v-ok{background:rgba(210,153,34,.15);color:var(--amber)}
  .v-weak{background:rgba(139,152,173,.12);color:var(--mute)}
  .v-no{background:rgba(248,81,73,.14);color:var(--neg)}
  .bdetails{margin-top:22px}
  .bdetails summary{cursor:pointer;color:var(--mute);font-family:var(--mono);font-size:.85rem;
    padding:6px 0}
  .bdetails summary:hover{color:var(--ink)}
  .bcard{cursor:pointer;transition:border-color .12s,transform .12s}
  .bcard:hover{border-color:var(--blue)}
  .bmore{font-size:.75rem;color:var(--blue);font-family:var(--mono);margin-top:12px}
  /* small "recent event report" box inside each card */
  .recbox{margin-top:13px;background:#111826;border:1px solid var(--grid);border-radius:8px;
    padding:10px 12px;cursor:pointer;transition:border-color .12s}
  .recbox:hover{border-color:var(--amber);background:#141c2c}
  .recl{font-family:var(--mono);font-size:.66rem;letter-spacing:.05em;color:var(--mute);
    text-transform:uppercase}
  .recv{font-family:var(--mono);font-size:.92rem;margin-top:4px;color:var(--ink)}
  .recmore{font-family:var(--mono);font-size:.68rem;color:var(--amber);margin-top:5px}
  /* "all past events" link + table rows */
  .histlink{margin-top:11px;font-family:var(--mono);font-size:.72rem;color:var(--blue);
    cursor:pointer;padding:8px 10px;border:1px dashed var(--grid);border-radius:8px;
    transition:border-color .12s,background .12s}
  .histlink:hover{border-color:var(--blue);background:#141c2c}
  .hrow{cursor:pointer;transition:background .1s}
  .hrow:hover{background:#1b2333}
  /* rankings tab */
  .rtabs{display:flex;gap:8px;flex-wrap:wrap;margin:8px 0 4px}
  .rtab{background:var(--panel);border:1px solid var(--grid);color:var(--mute);border-radius:8px;
    padding:8px 14px;cursor:pointer;font-weight:600;font-size:.9rem;font-family:inherit;transition:all .12s}
  .rtab:hover{color:var(--ink);border-color:var(--blue)}
  .rtab.active{background:var(--blue);color:#0b1220;border-color:var(--blue)}
  .rrow{cursor:pointer;transition:background .1s}
  .rrow:hover{background:#1b2333}
  .rk{font-family:var(--mono);font-size:1rem;width:44px;text-align:center}
  .rsym{font-weight:700;color:var(--ink)}
  .rnum{text-align:right;font-family:var(--mono);white-space:nowrap}
  .rbar{position:relative;height:12px;min-width:120px;background:#0e1420;border:1px solid var(--grid);
    border-radius:100px;overflow:hidden}
  .rbar>span{display:block;height:100%;border-radius:100px}
  /* screener */
  .scfilters{display:flex;gap:12px;flex-wrap:wrap;align-items:center;margin:8px 0 4px;
    background:var(--panel);border:1px solid var(--grid);border-radius:10px;padding:12px 14px}
  .scfilters label{font-size:.82rem;color:var(--mute);display:flex;align-items:center;gap:6px;font-family:var(--mono)}
  .scfilters input{width:70px;background:#0e1420;border:1px solid var(--grid);color:var(--ink);
    border-radius:6px;padding:5px 8px;font-family:var(--mono);font-size:.85rem}
  .scbtn{background:var(--blue);color:#0b1220;border:0;border-radius:8px;padding:7px 14px;
    font-weight:700;font-size:.85rem;cursor:pointer;font-family:inherit}
  .scbtn.ghost{background:transparent;border:1px solid var(--grid);color:var(--mute)}
  .scbtn:hover{filter:brightness(1.08)}
  .scH{cursor:pointer;user-select:none}.scH.rt{text-align:right}.scH:hover{color:var(--ink)}
  /* compare */
  .cmpbar{display:flex;gap:8px;flex-wrap:wrap;align-items:center;margin:8px 0}
  .cmpchip{background:var(--panel);border:1px solid var(--grid);border-radius:20px;padding:5px 12px;
    font-weight:600;font-size:.85rem}
  .cmpchip b{cursor:pointer;color:var(--mute);margin-left:4px}
  .cmpchip b:hover{color:var(--neg)}
  .cmpbar input{background:#0e1420;border:1px solid var(--grid);color:var(--ink);border-radius:8px;
    padding:6px 12px;font-family:inherit;font-size:.85rem}
  .cmphead{cursor:pointer;color:var(--blue)}.cmphead:hover{text-decoration:underline}
  .cmplab{color:var(--mute);font-size:.85rem;white-space:nowrap}
  .cmpsec td{background:var(--panel);color:var(--mute);font-family:var(--mono);font-size:.72rem;
    text-transform:uppercase;letter-spacing:.05em;font-weight:600}
  .hdate{color:var(--blue);font-weight:600}
  .hopen{opacity:.5;font-size:.8em}
  .hrow:hover .hopen{opacity:1}
  /* movable, interactive per-event graph panel */
  .evgraph{position:fixed;z-index:9999;width:460px;max-width:94vw;background:var(--panel);
    border:1px solid var(--grid);border-radius:12px;box-shadow:0 18px 50px rgba(0,0,0,.55);overflow:hidden}
  .evgraph-head{display:flex;justify-content:space-between;align-items:center;gap:10px;
    padding:9px 12px;background:#111826;border-bottom:1px solid var(--grid);cursor:move;
    font-weight:600;font-size:.86rem;color:var(--ink);user-select:none}
  .evgraph-head .evx{cursor:pointer;color:var(--mute);font-size:1rem;padding:0 4px}
  .evgraph-head .evx:hover{color:var(--neg,#f87171)}
  .evgraph-body{padding:12px 14px 4px;position:relative}
  .evtip{min-height:18px;text-align:center;font-family:var(--mono);font-size:.76rem;
    color:var(--ink);opacity:0;transition:opacity .1s;margin-top:2px}
  .evgraph-foot{display:flex;justify-content:space-between;gap:10px;flex-wrap:wrap;
    padding:8px 14px 12px;font-family:var(--mono);font-size:.72rem;color:var(--mute)}
  .evgraph-foot .evhint{opacity:.75}
  /* behaviour drill-in */
  .backbtn{background:var(--panel);border:1px solid var(--grid);color:var(--ink);border-radius:7px;
    padding:6px 13px;cursor:pointer;font-weight:600;font-size:.88rem;margin-bottom:12px}
  .backbtn:hover{background:#243044}
  .bhead2{font-size:1.35rem;font-weight:700;margin:2px 0 12px}
  .bhead2 .bn{font-size:.8rem;color:var(--mute);font-family:var(--mono);font-weight:400}
  .statrow{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-top:16px}
  @media(max-width:600px){.statrow{grid-template-columns:1fr 1fr}}
  .sbox{background:var(--panel);border:1px solid var(--grid);border-radius:10px;padding:13px 15px}
  .sbox .l{font-size:.68rem;text-transform:uppercase;color:var(--mute);font-family:var(--mono);
    letter-spacing:.05em}
  .sbox .v{font-size:1.3rem;font-weight:700;font-family:var(--mono);margin-top:4px}
  .banner{background:var(--panel);border:1px solid var(--grid);border-radius:10px;
    padding:22px;color:var(--mute);text-align:center;line-height:1.6}
  .banner b{color:var(--ink)}
  .dl{background:var(--blue);color:#0b1220;border:0;border-radius:8px;padding:10px 18px;
    font-weight:700;font-size:.95rem;cursor:pointer}
  .dl:hover{background:#7cb8ff}
  .dlbar{display:flex;align-items:center;gap:12px;flex-wrap:wrap;margin:0 0 12px}
  .dlbtn{background:var(--blue);color:#0b1220;border-radius:8px;padding:9px 16px;
    font-weight:700;font-size:.9rem;text-decoration:none}
  .dlbtn:hover{background:#7cb8ff}
  .dlnote{color:var(--mute);font-size:.8rem;font-family:var(--mono)}
  a{color:var(--blue)}
  .hint{color:var(--mute);font-size:.82rem;margin-top:16px;font-family:var(--mono)}

  /* ── new components: out-of-sample line + upcoming badges ── */
  .boos{font-family:var(--mono);font-size:.74rem;color:var(--mute);margin-top:6px}
  .boos b{font-weight:700}
  .ubadge{font-family:var(--mono);font-size:.6rem;font-weight:700;padding:.15em .55em;
    border-radius:20px;margin-left:6px;text-transform:uppercase;letter-spacing:.03em}
  .ubadge.conf{background:rgba(63,185,80,.15);color:var(--pos)}
  .ubadge.est{background:rgba(210,153,34,.16);color:var(--amber)}
  .uwhen{font-family:var(--mono);font-size:.66rem;color:var(--mute);margin-top:2px}

  /* ── visual polish (cosmetic only; no layout/logic change) ── */
  body{background:
    radial-gradient(1100px 520px at 15% -10%, #182444 0%, rgba(24,36,68,0) 60%),
    radial-gradient(900px 500px at 100% 0%, #10233a 0%, rgba(16,35,58,0) 55%),
    var(--bg) fixed;}
  .bar{background:rgba(15,19,27,.80);backdrop-filter:saturate(140%) blur(10px);
    -webkit-backdrop-filter:saturate(140%) blur(10px)}
  .bar h1{display:flex;align-items:center;gap:9px;letter-spacing:-.01em}
  .bar h1::before{content:"";width:11px;height:11px;border-radius:50%;flex:0 0 auto;
    background:linear-gradient(135deg,var(--blue),#7ee0a6);box-shadow:0 0 12px rgba(88,166,255,.55)}
  button{transition:background .13s,border-color .13s,transform .06s,box-shadow .13s}
  button:hover{border-color:#3a4a66}
  input{transition:border-color .13s,box-shadow .13s}
  input:focus{box-shadow:0 0 0 3px rgba(88,166,255,.18)}
  .tab{transition:color .13s,border-color .13s}
  .sym{background:linear-gradient(92deg,var(--ink) 30%,#93b4e8);-webkit-background-clip:text;
    background-clip:text;color:transparent}
  .card,.bcard,.chart,.banner,.tblwrap,.sbox{
    box-shadow:0 1px 2px rgba(0,0,0,.28),0 10px 26px rgba(0,0,0,.20)}
  .card{border-radius:12px;transition:transform .13s,border-color .13s,box-shadow .13s}
  .card:hover{transform:translateY(-2px);border-color:#3a4a66}
  .bcard{border-radius:14px}
  .bcard:hover{box-shadow:0 2px 4px rgba(0,0,0,.3),0 14px 34px rgba(0,0,0,.28)}
  .rrow{cursor:pointer;transition:background .1s}
  .rtab{transition:background .12s,color .12s,border-color .12s}
  .tblwrap{border-radius:12px}
  ::selection{background:rgba(88,166,255,.32)}
  /* tidy scrollbars (WebKit) */
  ::-webkit-scrollbar{width:11px;height:11px}
  ::-webkit-scrollbar-thumb{background:#2c374d;border-radius:8px;border:2px solid var(--bg)}
  ::-webkit-scrollbar-thumb:hover{background:#3a4a66}
</style>
</head>
<body>
  <div class="bar">
    <h1>NSE Stock Browser</h1>
    <button id="first">◀◀ First</button>
    <button id="prev">◀ Prev</button>
    <span class="searchwrap">
      <input id="search" placeholder="🔍 search symbol…" autocomplete="off">
      <div id="ac" class="ac"></div>
    </span>
    <button id="next">Next ▶</button>
    <button id="last">Last ▶▶</button>
    <span class="counter" id="counter">—</span>
  </div>

  <main>
    <div><span class="sym" id="sym">…</span><span class="have" id="have"></span>
         <span id="evstatus"></span></div>

    <div class="tabs" id="tabs">
      <button class="tab active" data-tab="fundamentals">Fundamentals</button>
      <button class="tab" data-tab="behaviour">📈 Behaviour</button>
      <button class="tab" data-tab="upcoming">📅 Upcoming</button>
      <button class="tab" data-tab="rankings">🏆 Rankings</button>
      <button class="tab" data-tab="screener">🔎 Screener</button>
      <button class="tab" data-tab="compare">⚖ Compare</button>
      <button class="tab" data-tab="announcements">Announcements</button>
      <button class="tab" data-tab="board_meetings">Board Meetings</button>
      <button class="tab" data-tab="corporate_actions">Corporate Actions</button>
      <button class="tab" data-tab="financial_results">Financial Results</button>
    </div>

    <!-- Fundamentals view -->
    <div id="fundview">
      <div class="cards" id="cards"></div>
      <div class="charts">
        <div class="chart">
          <span class="cs">Sales vs Net Profit</span><span class="ct">Annual trend (₹ cr)</span>
          <canvas id="c_annual"></canvas>
        </div>
        <div class="chart">
          <span class="cs">green = YoY up</span><span class="ct">Quarterly Net Profit (₹ cr)</span>
          <canvas id="c_quarter"></canvas>
        </div>
      </div>
    </div>

    <!-- Behaviour view -->
    <div id="behaviourview" style="display:none"></div>

    <!-- Upcoming events watchlist (events still ahead + historical edge) -->
    <div id="upcomingview" style="display:none"></div>

    <!-- Rankings view (all stocks ordered by win rate, per event) -->
    <div id="rankview" style="display:none"></div>

    <!-- Screener view (filter all stocks by fundamentals) -->
    <div id="screenview" style="display:none"></div>

    <!-- Compare view (2-3 stocks side by side) -->
    <div id="compareview" style="display:none"></div>

    <!-- Event view -->
    <div id="eventsview" style="display:none"></div>

    <div class="hint">🔍 search any of 2,363 stocks · ← / → arrow keys to step · click a tab for event details.</div>
  </main>

<script>
const COL={ink:"#e6edf3",mute:"#8b98ad",grid:"#2a3346",panel:"#171d2b",
           blue:"#58a6ff",amber:"#d29922",pos:"#3fb950",neg:"#f85149"};
let SYMS=[], idx=0, activeTab="fundamentals", stock=null, eventsCache={}, READY=new Set();

/*__DATA_SOURCE__*/           /* swapped by build_static.py for the GitHub Pages build */
const STATIC=false;
const SYMBOLS_URL="/api/symbols";
function STOCK_URL(s){return "/api/stock?sym="+encodeURIComponent(s);}
function EVENTS_URL(s){return "/api/events?sym="+encodeURIComponent(s);}
/*__END_DATA_SOURCE__*/

const $=id=>document.getElementById(id);
function fmt(v,suf="",nd=0){ return (v===null||v===undefined)?"—":
  Number(v).toLocaleString("en-IN",{minimumFractionDigits:nd,maximumFractionDigits:nd})+suf; }
function esc(s){ return String(s).replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;"); }

/* ---------- canvas charts (HiDPI aware) ---------- */
function setup(cv){
  const r=cv.getBoundingClientRect(), dpr=window.devicePixelRatio||1;
  cv.width=r.width*dpr; cv.height=r.height*dpr;
  const ctx=cv.getContext("2d"); ctx.scale(dpr,dpr);
  return {ctx,w:r.width,h:r.height};
}
const M={l:56,r:18,t:14,b:30};
function niceBounds(vals,incZero){
  let v=vals.filter(x=>x!==null&&!isNaN(x)); if(incZero)v=v.concat([0]);
  if(!v.length)return null;
  let lo=Math.min(...v),hi=Math.max(...v),pad=((hi-lo)||1)*0.12;
  return {lo:lo-pad,hi:hi+pad};
}
function hexA(hex,a){                       // #rrggbb -> rgba()
  const c=hex.replace("#",""); const n=parseInt(c.length===3?c.replace(/(.)/g,"$1$1"):c,16);
  return `rgba(${(n>>16)&255},${(n>>8)&255},${n&255},${a})`;
}
function rrect(ctx,x,y,w,hh,r){
  r=Math.min(r,Math.abs(w)/2,Math.abs(hh)/2);
  if(ctx.roundRect){ctx.beginPath();ctx.roundRect(x,y,w,hh,r);return;}
  ctx.beginPath();ctx.moveTo(x+r,y);ctx.arcTo(x+w,y,x+w,y+hh,r);ctx.arcTo(x+w,y+hh,x,y+hh,r);
  ctx.arcTo(x,y+hh,x,y,r);ctx.arcTo(x,y,x+w,y,r);ctx.closePath();
}
function yaxis(ctx,w,h,lo,hi){
  ctx.font="11px Consolas"; ctx.textAlign="right"; ctx.textBaseline="middle";
  for(let i=0;i<=5;i++){
    const val=lo+(hi-lo)*i/5, y=h-M.b-(val-lo)/((hi-lo)||1)*(h-M.b-M.t);
    ctx.strokeStyle="rgba(138,152,173,.13)"; ctx.lineWidth=1; ctx.setLineDash([3,4]);
    ctx.beginPath();ctx.moveTo(M.l,y);ctx.lineTo(w-M.r,y);ctx.stroke(); ctx.setLineDash([]);
    ctx.fillStyle=COL.mute; ctx.fillText(Math.round(val).toLocaleString("en-IN"),M.l-8,y);
  }
}
function xlab(ctx,w,h,labels){
  ctx.font="10px Consolas"; ctx.fillStyle=COL.mute; ctx.textAlign="center"; ctx.textBaseline="top";
  const n=labels.length, iw=w-M.l-M.r;
  labels.forEach((L,i)=>{const x=M.l+(n<2?iw/2:iw*i/(n-1)); ctx.fillText(L,x,h-M.b+8);});
}
function empty(ctx,w,h,msg){ ctx.fillStyle=COL.mute;ctx.font="13px sans-serif";
  ctx.textAlign="center";ctx.textBaseline="middle";ctx.fillText(msg||"no data",w/2,h/2); }

function lineChart(cv,labels,seriesList){
  const {ctx,w,h}=setup(cv); ctx.clearRect(0,0,w,h);
  const all=seriesList.flatMap(s=>s.data);
  const b=niceBounds(all,false); if(!b||!labels.length){empty(ctx,w,h,"no annual P&L");return;}
  yaxis(ctx,w,h,b.lo,b.hi); xlab(ctx,w,h,labels);
  const n=labels.length,iw=w-M.l-M.r,base=h-M.b;
  const X=i=>M.l+(n<2?iw/2:iw*i/(n-1));
  const Y=v=>h-M.b-(v-b.lo)/((b.hi-b.lo)||1)*(h-M.b-M.t);
  seriesList.forEach((s,si)=>{
    const pts=[]; s.data.forEach((v,i)=>{ if(v!==null&&!isNaN(v)) pts.push([X(i),Y(v)]); });
    if(!pts.length)return;
    // soft gradient area under the line
    const g=ctx.createLinearGradient(0,M.t,0,base);
    g.addColorStop(0,hexA(s.color,.30)); g.addColorStop(1,hexA(s.color,0));
    ctx.beginPath();ctx.moveTo(pts[0][0],base);
    pts.forEach(p=>ctx.lineTo(p[0],p[1])); ctx.lineTo(pts[pts.length-1][0],base);
    ctx.closePath(); ctx.fillStyle=g; ctx.fill();
    // glowing smooth line
    ctx.save(); ctx.shadowColor=hexA(s.color,.55); ctx.shadowBlur=9;
    ctx.strokeStyle=s.color; ctx.lineWidth=2.5; ctx.lineJoin="round"; ctx.lineCap="round";
    ctx.beginPath(); pts.forEach((p,i)=>i?ctx.lineTo(p[0],p[1]):ctx.moveTo(p[0],p[1])); ctx.stroke();
    ctx.restore();
    // ringed dots
    pts.forEach(p=>{ ctx.fillStyle=COL.panel; ctx.beginPath();ctx.arc(p[0],p[1],4.5,0,7);ctx.fill();
      ctx.fillStyle=s.color; ctx.beginPath();ctx.arc(p[0],p[1],2.6,0,7);ctx.fill(); });
    // legend chip
    ctx.fillStyle=COL.ink;ctx.font="600 11px sans-serif";ctx.textAlign="left";ctx.textBaseline="middle";
    const ly=M.t+8+si*17;
    ctx.strokeStyle=s.color;ctx.lineWidth=3;ctx.lineCap="round";ctx.beginPath();
    ctx.moveTo(w-M.r-118,ly);ctx.lineTo(w-M.r-98,ly);ctx.stroke();
    ctx.fillText(s.name,w-M.r-92,ly);
  });
}
function barChart(cv,labels,vals,colors){
  const {ctx,w,h}=setup(cv); ctx.clearRect(0,0,w,h);
  const b=niceBounds(vals,true); if(!b||!labels.length){empty(ctx,w,h,"no quarterly data");return;}
  yaxis(ctx,w,h,b.lo,b.hi); xlab(ctx,w,h,labels);
  const n=labels.length,iw=w-M.l-M.r,slot=iw/n,bw=Math.min(slot*0.62,34);
  const Y=v=>h-M.b-(v-b.lo)/((b.hi-b.lo)||1)*(h-M.b-M.t);
  const yz=Y(0);
  ctx.strokeStyle="rgba(138,152,173,.35)";ctx.lineWidth=1;
  ctx.beginPath();ctx.moveTo(M.l,yz);ctx.lineTo(w-M.r,yz);ctx.stroke();
  vals.forEach((v,i)=>{ if(v===null||isNaN(v))return;
    const cx=M.l+slot*(i+0.5),y=Y(v),col=colors[i]||COL.blue;
    const top=Math.min(yz,y),hh=Math.abs(yz-y);
    const g=ctx.createLinearGradient(0,top,0,top+hh);
    if(v>=0){g.addColorStop(0,hexA(col,.95));g.addColorStop(1,hexA(col,.55));}
    else{g.addColorStop(0,hexA(col,.55));g.addColorStop(1,hexA(col,.95));}
    ctx.fillStyle=g; rrect(ctx,cx-bw/2,top,bw,Math.max(hh,1),4); ctx.fill();
    ctx.fillStyle=COL.ink;ctx.font="600 10px Consolas";ctx.textAlign="center";
    ctx.textBaseline=v>=0?"bottom":"top";
    ctx.fillText(Math.round(v).toLocaleString("en-IN"),cx,v>=0?y-4:y+4);
  });
}

/* ---------- data ---------- */
async function loadStock(){
  const sym=SYMS[idx];
  $("counter").textContent=(idx+1).toLocaleString()+" / "+SYMS.length.toLocaleString();
  $("search").value=sym;
  stock=await (await fetch(STOCK_URL(sym))).json();
  $("sym").textContent=stock.symbol;
  $("have").textContent="statements: "+(stock.have.join(", ")||"none");
  updateEvStatus();
  showTab(activeTab);
}
function updateEvStatus(){
  const on=READY.has(SYMS[idx]);
  $("evstatus").innerHTML=on
    ? '<span class="rdy">✓ event feeds downloaded</span>'
    : '<span class="notrdy">○ event feeds not downloaded</span>';
}
async function ensureEvents(sym){
  if(!eventsCache[sym]){
    try{
      const r=await fetch(EVENTS_URL(sym));
      eventsCache[sym]= r.ok ? await r.json() : {symbol:sym,available:false,feeds:{}};
    }catch(e){ eventsCache[sym]={symbol:sym,available:false,feeds:{}}; }
  }
  return eventsCache[sym];
}

/* ---------- rendering ---------- */
function renderFundamentals(){
  const c=stock.cards;
  const npCol=(c.net_profit!==null&&c.net_profit<0)?COL.neg:COL.pos;
  $("cards").innerHTML=[
    card("Sales (latest FY)",fmt(c.sales," cr"),fmt(c.sales_g,"% growth",1),COL.ink),
    card("Net Profit",fmt(c.net_profit," cr"),fmt(c.net_profit_g,"% growth",1),npCol),
    card("Operating margin",fmt(c.opm,"%",1),"OPM",COL.blue),
    card("ROCE",fmt(c.roce,"%",1),"Debt/Equity "+fmt(c.de,"",2),COL.amber),
  ].join("");
  lineChart($("c_annual"),stock.annual.labels,[
    {name:"Sales",color:COL.blue,data:stock.annual.sales},
    {name:"Net Profit",color:COL.amber,data:stock.annual.net_profit},
  ]);
  const qc=stock.quarterly.net_profit.map((_,i)=>{
    const g=stock.quarterly.yoy[i]; return (g!==undefined&&g!==null&&g>=0)?COL.pos:COL.neg; });
  barChart($("c_quarter"),stock.quarterly.labels,stock.quarterly.net_profit,qc);
}
function card(t,v,s,col){
  return `<div class="card"><div class="t">${t}</div>`+
         `<div class="v" style="color:${col}">${v}</div><div class="s">${s}</div></div>`;
}

async function renderEvents(feed){
  const ev=$("eventsview");
  ev.innerHTML='<div class="enote">loading…</div>';
  const pay=await ensureEvents(SYMS[idx]);
  if(feed!==activeTab)return;                    // stale (user moved on)
  if(!pay.available){
    ev.innerHTML=`<div class="banner">No event feeds for <b>${esc(pay.symbol)}</b> yet.<br>`+
      `Announcements, board meetings, corporate actions and financial results can be pulled `+
      `live from NSE.<br><br>`+
      `<button class="dl" onclick="fetchNSE()">⬇ Download NSE feeds for ${esc(pay.symbol)}</button></div>`;
    return;
  }
  const F=pay.feeds[feed];
  if(!F||!F.rows.length){
    ev.innerHTML=`<div class="banner">No <b>${esc((F&&F.label)||feed)}</b> rows for <b>${esc(pay.symbol)}</b>.<br><br>`+
      `<button class="dl" onclick="fetchNSE()">⬇ Re-download NSE feeds for ${esc(pay.symbol)}</button></div>`;
    return;
  }
  // Financial Results: offer a full statements workbook built from the LOCAL
  // folders (pnl/balance_sheet/cash_flow/ratios/quarterly) — no online fetch.
  let html='';
  if(feed==='financial_results'){
    const su="/api/statements_excel?sym="+encodeURIComponent(SYMS[idx]);
    html+=`<div class="dlbar"><a class="dlbtn" href="${su}" `+
      `title="P&L, Balance Sheet, Cash Flow, Ratios & Quarterly from local data">`+
      `⬇ Download Excel (financial statements)</a>`+
      `<span class="dlnote">built from your offline folders — the 📊 XBRL links below are raw NSE filings</span></div>`;
  }
  html+='<div class="tblwrap"><table><thead><tr>'+
    F.columns.map(c=>`<th>${esc(c.label)}</th>`).join('')+'</tr></thead><tbody>';
  F.rows.forEach(r=>{
    html+='<tr>'+F.columns.map(c=>{
      const v=r[c.key]||'';
      if(c.type==='link'){
        if(!v.startsWith('http')) return `<td>${esc(v)}</td>`;
        if(c.key==='xbrl'){        // Download Excel -> built from the LOCAL folders
          // the static (GitHub Pages) build has no server, so link to the NSE doc instead
          if(STATIC) return `<td><a href="${v}" target="_blank" rel="noopener" title="${esc(v)}">📄 XBRL ↗</a></td>`;
          const u="/api/statements_excel?sym="+encodeURIComponent(SYMS[idx]);
          return `<td><a href="${u}" title="Download financial statements from local data (P&L, Balance Sheet, Cash Flow, Ratios, Quarterly)">📊 Excel ↓</a></td>`;
        }
        // attachments: on the live app we proxy through the server (NSE blocks
        // some direct hits); on the static build there is no server, so link direct.
        const pdf=v.toLowerCase().endsWith(".pdf");
        const f=STATIC ? v : "/api/file?url="+encodeURIComponent(v);
        return `<td><a href="${f}" target="_blank" rel="noopener" title="${esc(v)}">`+
               `${pdf?"📄 PDF ↗":"📄 view ↗"}</a></td>`;
      }
      if(c.key==='desc') return `<td class="cat">${esc(v)}</td>`;
      return `<td>${esc(v)}</td>`;
    }).join('')+'</tr>';
  });
  html+='</tbody></table></div>';
  if(F.total>F.shown) html+=`<div class="enote">Showing latest ${F.shown} of ${F.total.toLocaleString()} rows.</div>`;
  else html+=`<div class="enote">${F.total.toLocaleString()} row(s).</div>`;
  ev.innerHTML=html;
}

async function fetchNSE(){
  const sym=SYMS[idx];
  const ev=$("eventsview");
  if(STATIC){
    ev.innerHTML=`<div class="banner">This is a <b>static snapshot</b> — live NSE download isn't available here.<br>`+
      `To pull <b>${esc(sym)}</b>'s feeds from NSE, run the app locally:<br><br>`+
      `<span class="enote">python stock_server.py</span></div>`;
    return;
  }
  ev.innerHTML=`<div class="banner">Downloading NSE feeds for <b>${esc(sym)}</b>…<br>`+
    `<span class="enote">(announcements can be a few thousand rows — give it 5–15 s)</span></div>`;
  try{
    const j=await (await fetch("/api/fetch?sym="+encodeURIComponent(sym))).json();
    if(!j.ok){
      const reason=j.error || (j.errors && Object.values(j.errors)[0]) || "NSE refused the request";
      ev.innerHTML=`<div class="banner">Couldn't download <b>${esc(sym)}</b> from NSE.<br>`+
        `<span class="enote">${esc(reason)}</span><br><br>`+
        `NSE rate-limits rapid requests. Wait ~30–60 s, then:<br><br>`+
        `<button class="dl" onclick="fetchNSE()">Retry</button></div>`;
      return;
    }
    READY.add(sym);
    updateEvStatus();
    delete eventsCache[sym];
    renderEvents(activeTab);
  }catch(e){
    ev.innerHTML=`<div class="banner">Download error: ${esc(e)}<br><br>`+
      `NSE rate-limits rapid requests — wait ~30–60 s and retry.<br><br>`+
      `<button class="dl" onclick="fetchNSE()">Retry</button></div>`;
  }
}

let behaviourCache={}, behaviourData=null;
async function renderBehaviour(){
  const bv=$("behaviourview"), sym=SYMS[idx];
  bv.innerHTML='<div class="enote">computing how '+esc(sym)+' moved around its events…</div>';
  let d;
  try{
    if(!behaviourCache[sym]) behaviourCache[sym]=await (await fetch("/api/behaviour?sym="+encodeURIComponent(sym))).json();
    d=behaviourCache[sym];
  }catch(e){ bv.innerHTML='<div class="banner">Error computing behaviour: '+esc(e)+'</div>'; return; }
  if(activeTab!=="behaviour")return;
  if(d.error){ bv.innerHTML=`<div class="banner">Couldn't analyse <b>${esc(sym)}</b>.<br><span class="enote">${esc(d.error)}</span></div>`; return; }
  if(!d.available){
    bv.innerHTML=`<div class="banner">Can't analyse <b>${esc(sym)}</b> yet — ${esc(d.reason||'no event feeds')}.<br><br>`+
      `Download its NSE event feeds first (any event tab → ⬇ Download).</div>`;
    return;
  }
  const order=["RESULTS","BOARD_MEETING","CORPORATE_ACTION","ANNOUNCEMENT"];
  const keys=order.filter(k=>d.types[k]).concat(Object.keys(d.types).filter(k=>!order.includes(k)));

  behaviourData=d;
  let html='<div class="bintro">The best time to buy <b>'+esc(sym)+'</b> around each kind of event — '+
    'based on what actually happened in the past. <b>Click any card</b> for its graph &amp; full summary.</div><div class="bcards">';
  for(const et of keys){
    const t=d.types[et], b=t.best;
    const [vc,vt]=verdict(b.win_rate_pct,b.avg_return_pct);
    const barcol=b.win_rate_pct>=55?COL.pos:(b.win_rate_pct>=50?COL.amber:COL.neg);
    const ps=b.avg_return_pct>=0?'+':'';
    const dbf=b.days_before, daf=b.days_after;
    let rec='';
    if(t.last){
      const L=t.last, e=L.days.indexOf(0);
      const p0=L.prices[0], pe=L.prices[e], pa=L.prices[L.prices.length-1];
      const mv=(pa/p0-1)*100;
      rec=`<div class="recbox" onclick="event.stopPropagation();openRecent('${et}')">
        <div class="recl">📋 Recent event report · ${esc(L.date)}</div>
        <div class="recv">₹${rup(p0)} → ₹${rup(pe)} → ₹${rup(pa)}
          &nbsp;<b style="color:${mv>=0?COL.pos:COL.neg}">${mv>=0?'+':''}${mv.toFixed(2)}%</b></div>
        <div class="recmore">click for full report →</div></div>`;
    }
    html+=`<div class="bcard" onclick="openBeh('${et}')">
      <div class="bcard-top"><span class="bet">${esc(pretty(et))}</span><span class="bn">${t.n_events} past events</span></div>
      <div class="brec">Buy <b>${dbf} day${dbf>1?'s':''} before</b> the event, then sell <b>${daf} day${daf>1?'s':''} after</b>.</div>
      <div class="bstats">
        <div class="bstat">
          <div class="blabel">Chance of profit (win rate)</div>
          <div class="bwinbar"><span style="width:${Math.max(2,Math.min(100,b.win_rate_pct))}%;background:${barcol}"></span></div>
          <div class="bwinval" style="color:${barcol}">${b.win_rate_pct}%</div>
        </div>
        <div class="bstat">
          <div class="blabel">Average gain per trade</div>
          <div class="bprofit" style="color:${b.avg_return_pct>=0?COL.pos:COL.neg}">${ps}${b.avg_return_pct}%</div>
        </div>
      </div>
      <div class="bverdict v-${vc}">${vt}</div>
      <div class="bmore">📈 click for graph &amp; performance summary →</div>
      ${rec}
      ${(t.history&&t.history.length)?`<div class="histlink" onclick="event.stopPropagation();openHistory('${et}')">📜 all ${t.history.length} past events — what happened each time →</div>`:''}
    </div>`;
  }
  html+='</div>';
  bv.innerHTML=html;
}

/* ---- the recent-event report (opened from the small box on each card) ---- */
function openRecent(et){
  const d=behaviourData; if(!d||!d.types[et]||!d.types[et].last)return;
  const t=d.types[et], L=t.last, bv=$("behaviourview");
  const e=L.days.indexOf(0), n=L.prices.length;
  const p0=L.prices[0], pe=L.prices[e], pa=L.prices[n-1];
  const m1=(pe/p0-1)*100, m2=(pa/pe-1)*100, m3=(pa/p0-1)*100;
  const c=v=>v>=0?COL.pos:COL.neg, s=v=>(v>=0?'+':'')+v.toFixed(2)+'%';
  const rows=L.days.map((dd,i)=>{
    const pct=(L.prices[i]/pe-1)*100;
    const lab=dd===0?'EVENT DAY':(dd>0?'T+'+dd:'T'+dd);
    const key=(i===0||dd===0||i===n-1);
    return `<tr${key?' style="background:#1b2333;font-weight:700"':''}>`+
      `<td>${lab}</td><td style="text-align:right">₹${rup(L.prices[i])}</td>`+
      `<td style="text-align:right;color:${c(pct)}">${s(pct)}</td></tr>`;
  }).join('');
  const outcome=m3>=0?'PROFIT':'LOSS';
  bv.innerHTML=`
    <button class="backbtn" onclick="renderBehaviour()">← back to all events</button>
    <div class="bhead2">Recent ${esc(pretty(et))} — full report
      <span class="bn">· ${esc(d.symbol)} · ${esc(L.date)}</span></div>
    <div class="bverdict v-${m3>=0?'strong':'no'}" style="display:inline-block">
      ${outcome}: buying 6 days before and selling 3 days after would have returned ${s(m3)}</div>
    <div class="statrow" style="grid-template-columns:repeat(3,1fr);margin-top:14px">
      <div class="sbox"><div class="l">Price 6 days before</div><div class="v" style="color:${COL.amber}">₹${rup(p0)}</div></div>
      <div class="sbox"><div class="l">Price on event day</div><div class="v">₹${rup(pe)}</div></div>
      <div class="sbox"><div class="l">Price 3 days after</div><div class="v" style="color:${COL.pos}">₹${rup(pa)}</div></div>
    </div>
    <div class="statrow" style="grid-template-columns:repeat(3,1fr)">
      <div class="sbox"><div class="l">Before → Event</div><div class="v" style="color:${c(m1)}">${s(m1)}</div></div>
      <div class="sbox"><div class="l">Event → After</div><div class="v" style="color:${c(m2)}">${s(m2)}</div></div>
      <div class="sbox"><div class="l">Total move</div><div class="v" style="color:${c(m3)}">${s(m3)}</div></div>
    </div>
    <div class="chart" style="margin-top:14px">
      <span class="cs">move cursor 👆</span>
      <span class="ct">Price around this event (₹)</span>
      <canvas id="bmeter"></canvas>
    </div>
    <div class="bhead" style="margin-top:18px">Day-by-day prices</div>
    <div class="tblwrap"><table><thead><tr><th>Day</th>
      <th style="text-align:right">Price</th><th style="text-align:right">vs event day</th>
      </tr></thead><tbody>${rows}</tbody></table></div>`;
  resultMeter($("bmeter"),L);
}

/* ---- ALL past events of one type: what happened each time ---- */
function openHistory(et){
  const d=behaviourData; if(!d||!d.types[et]||!d.types[et].history)return;
  const t=d.types[et], H=t.history, bv=$("behaviourview");
  const c=v=>v>=0?COL.pos:COL.neg, s=v=>(v>=0?'+':'')+v.toFixed(2)+'%';
  const wins=H.filter(e=>e.total_pct>0).length;
  const wr=H.length?Math.round(wins/H.length*100):0;
  const avg=H.length?H.reduce((a,e)=>a+e.total_pct,0)/H.length:0;
  const best=H.reduce((a,e)=>e.total_pct>a?e.total_pct:a,-1e9);
  const worst=H.reduce((a,e)=>e.total_pct<a?e.total_pct:a,1e9);
  window._histET=pretty(et); window._histSym=d.symbol;
  const rows=H.map((e,i)=>{
    const out=e.total_pct>=0?'WIN':'LOSS';
    return `<tr class="hrow" onclick="openEventGraph(${i})" title="click to open a movable price graph for this event">
      <td class="hdate">${esc(e.date)} <span class="hopen">📈</span></td>
      <td style="text-align:right">₹${rup(e.buy)}</td>
      <td style="text-align:right">₹${rup(e.event)}</td>
      <td style="text-align:right">₹${rup(e.sell)}</td>
      <td style="text-align:right;color:${c(e.before_pct)}">${s(e.before_pct)}</td>
      <td style="text-align:right;color:${c(e.after_pct)}">${s(e.after_pct)}</td>
      <td style="text-align:right;color:${c(e.total_pct)};font-weight:700">${s(e.total_pct)}</td>
      <td style="text-align:center;color:${c(e.total_pct)};font-weight:700">${out}</td></tr>`;
  }).join('');
  window._histData=H;
  bv.innerHTML=`
    <button class="backbtn" onclick="renderBehaviour()">← back to all events</button>
    <div class="bhead2">All past ${esc(pretty(et))} — what happened each time
      <span class="bn">· ${esc(d.symbol)} · ${H.length} events</span></div>
    <div class="statrow" style="grid-template-columns:repeat(5,1fr);margin-top:12px">
      <div class="sbox"><div class="l">Past events</div><div class="v">${H.length}</div></div>
      <div class="sbox"><div class="l">Win rate</div><div class="v" style="color:${wr>=55?COL.pos:(wr>=50?COL.amber:COL.neg)}">${wr}%</div></div>
      <div class="sbox"><div class="l">Average total move</div><div class="v" style="color:${c(avg)}">${s(avg)}</div></div>
      <div class="sbox"><div class="l">Best</div><div class="v" style="color:${COL.pos}">${s(best)}</div></div>
      <div class="sbox"><div class="l">Worst</div><div class="v" style="color:${COL.neg}">${s(worst)}</div></div>
    </div>
    <div class="enote">Each row: buy 6 trading days before the event, sell 3 days after — the same
      window as the “recent event report”, shown for every past event (newest first). Click a row for its price chart.</div>
    <div id="histchart"></div>
    <div class="tblwrap" style="margin-top:12px"><table><thead><tr>
      <th>Event date</th><th style="text-align:right">Buy (T-6)</th>
      <th style="text-align:right">Event day</th><th style="text-align:right">Sell (T+3)</th>
      <th style="text-align:right">Before→Event</th><th style="text-align:right">Event→After</th>
      <th style="text-align:right">Total</th><th style="text-align:center">Result</th>
      </tr></thead><tbody>${rows}</tbody></table></div>`;
}
function histRow(i){
  const e=(window._histData||[])[i]; if(!e)return;
  const box=$("histchart");
  box.innerHTML=`<div class="chart" style="margin-top:14px">
    <span class="cs">move cursor 👆</span>
    <span class="ct">Price around the ${esc(e.date)} event (₹)</span>
    <canvas id="hmeter"></canvas></div>`;
  resultMeter($("hmeter"),{date:e.date,days:e.days,prices:e.prices});
  box.scrollIntoView({behavior:"smooth",block:"nearest"});
}

/* ---- a MOVABLE, interactive price graph for one past event ---- */
function openEventGraph(i){
  const H=window._histData||[], e=H[i]; if(!e)return;
  const et=window._histET||'Event', sym=window._histSym||'';
  const old=document.getElementById('evgraph'); if(old) old.remove();

  const days=e.days, px=e.prices, n=px.length, ei=days.indexOf(0);
  const W=460,Hh=260,L=52,R=16,T=20,B=36, pw=W-L-R, ph=Hh-T-B;
  const lo=Math.min(...px), hi=Math.max(...px), pad=(hi-lo)*0.12||1;
  const vmin=lo-pad, vmax=hi+pad;
  const X=k=>L+k/(n-1)*pw, Y=v=>T+(vmax-v)/(vmax-vmin)*ph;
  const P=v=>((v/px[ei]-1)*100), sp=v=>(v>=0?'+':'')+v.toFixed(2)+'%';
  const BL='#5b9bff', GR='#2b3446', INK='#e6e9ef', MU='#98a2b3', PO='#5ec98a', NE='#f87171';

  let g=`<svg id="evsvg" viewBox="0 0 ${W} ${Hh}" style="width:100%;height:auto;display:block">`;
  // gridlines + price ticks
  for(let t=0;t<=3;t++){ const v=vmin+(vmax-vmin)*t/3, y=Y(v);
    g+=`<line x1="${L}" y1="${y.toFixed(1)}" x2="${W-R}" y2="${y.toFixed(1)}" stroke="${GR}" stroke-width="1"/>`;
    g+=`<text x="${L-8}" y="${(y+4).toFixed(1)}" fill="${MU}" font-size="10" text-anchor="end">₹${Math.round(v)}</text>`; }
  // event-day marker
  g+=`<line x1="${X(ei).toFixed(1)}" y1="${T}" x2="${X(ei).toFixed(1)}" y2="${T+ph}" stroke="${MU}" stroke-width="1.2" stroke-dasharray="4 4"/>`;
  g+=`<text x="${X(ei).toFixed(1)}" y="${T-6}" fill="${MU}" font-size="10" text-anchor="middle">EVENT</text>`;
  // x labels
  for(let k=0;k<n;k++){ const d=days[k]; if(d%2===0||d===0)
    g+=`<text x="${X(k).toFixed(1)}" y="${Hh-18}" fill="${MU}" font-size="9.5" text-anchor="middle">${d===0?'T':'T'+(d>0?'+':'')+d}</text>`; }
  // price line + points
  let pts=px.map((v,k)=>`${X(k).toFixed(1)},${Y(v).toFixed(1)}`).join(' ');
  g+=`<polyline points="${pts}" fill="none" stroke="${BL}" stroke-width="2.2" stroke-linejoin="round"/>`;
  px.forEach((v,k)=>{ const col=(k===0)?PO:(k===n-1)?NE:BL;
    g+=`<circle cx="${X(k).toFixed(1)}" cy="${Y(v).toFixed(1)}" r="${(k===0||k===n-1)?4:2.6}" fill="${col}"/>`; });
  // crosshair (updated on hover)
  g+=`<line id="evx" x1="0" y1="${T}" x2="0" y2="${T+ph}" stroke="${BL}" stroke-width="1" opacity="0"/>`;
  g+=`<circle id="evdot" cx="0" cy="0" r="5" fill="#fff" stroke="${BL}" stroke-width="2.5" opacity="0"/>`;
  g+=`<rect id="evhit" x="${L}" y="${T}" width="${pw}" height="${ph}" fill="transparent"/>`;
  g+=`</svg>`;

  const el=document.createElement('div');
  el.id='evgraph'; el.className='evgraph';
  el.style.left=Math.max(20,(window.innerWidth-W-40)/2)+'px';
  el.style.top='120px';
  el.innerHTML=`
    <div class="evgraph-head" id="evhead">
      <span>📈 ${esc(et)} · ${esc(sym)} · ${esc(e.date)}</span>
      <span class="evx" title="close" onclick="document.getElementById('evgraph').remove()">✕</span>
    </div>
    <div class="evgraph-body">${g}<div class="evtip" id="evtip"></div></div>
    <div class="evgraph-foot">
      <span>buy T-6 ₹${rup(e.buy)} · sell T+3 ₹${rup(e.sell)} · total <b style="color:${e.total_pct>=0?PO:NE}">${sp(e.total_pct)}</b></span>
      <span class="evhint">drag title to move · hover line for values</span></div>`;
  document.body.appendChild(el);

  // hover crosshair
  const svg=$("evsvg"), xln=$("evx"), dot=$("evdot"), tip=$("evtip"), hit=$("evhit");
  function move(ev){
    const r=svg.getBoundingClientRect(), sx=(ev.clientX-r.left)/r.width*W;
    let k=Math.round((sx-L)/pw*(n-1)); k=Math.max(0,Math.min(n-1,k));
    const x=X(k), y=Y(px[k]);
    xln.setAttribute('x1',x); xln.setAttribute('x2',x); xln.setAttribute('opacity','.6');
    dot.setAttribute('cx',x); dot.setAttribute('cy',y); dot.setAttribute('opacity','1');
    const d=days[k], lab=d===0?'EVENT DAY':('T'+(d>0?'+':'')+d);
    tip.style.opacity=1;
    tip.innerHTML=`<b>${lab}</b> · ₹${rup(px[k])} · <span style="color:${P(px[k])>=0?PO:NE}">${sp(P(px[k]))}</span> vs event`;
  }
  hit.addEventListener('mousemove',move);
  svg.addEventListener('mouseleave',()=>{xln.setAttribute('opacity','0');dot.setAttribute('opacity','0');tip.style.opacity=0;});

  // drag to move the whole panel
  const head=$("evhead"); let dx=0,dy=0,drag=false;
  head.addEventListener('mousedown',ev=>{drag=true;dx=ev.clientX-el.offsetLeft;dy=ev.clientY-el.offsetTop;ev.preventDefault();});
  document.addEventListener('mousemove',ev=>{ if(!drag)return;
    el.style.left=Math.max(0,Math.min(window.innerWidth-60,ev.clientX-dx))+'px';
    el.style.top=Math.max(0,Math.min(window.innerHeight-40,ev.clientY-dy))+'px'; });
  document.addEventListener('mouseup',()=>{drag=false;});
}

/* ---- drill-in: one event type, with graph + performance summary ---- */
function openBeh(et){
  const d=behaviourData; if(!d||!d.types[et])return;
  const t=d.types[et], b=t.best, bv=$("behaviourview");
  const [vc,vt]=verdict(b.win_rate_pct,b.avg_return_pct);
  const pc=v=>(v==null?'—':(v>=0?'+':'')+v+'%');
  const col=v=>(v==null?COL.mute:(v>=0?COL.pos:COL.neg));
  bv.innerHTML=`
    <button class="backbtn" onclick="renderBehaviour()">← back to all events</button>
    <div class="bhead2">${esc(pretty(et))} <span class="bn">· ${esc(d.symbol)} · ${t.n_events} past events</span></div>
    <div class="bverdict v-${vc}" style="display:inline-block">${vt}</div>
    <div class="brec" style="margin-top:12px">Best rule: buy <b>${b.days_before} day${b.days_before>1?'s':''} before</b> the event, sell <b>${b.days_after} day${b.days_after>1?'s':''} after</b>.</div>
    <div class="chart" style="margin-top:8px">
      <span class="cs">🟢 BUY · 🔴 SELL · move cursor 👆</span>
      <span class="ct">How the price usually moves around a ${esc(pretty(et)).toLowerCase()}</span>
      <canvas id="bpath"></canvas>
    </div>
    <div class="statrow">
      <div class="sbox"><div class="l">Chance of profit</div><div class="v" style="color:${b.win_rate_pct>=55?COL.pos:(b.win_rate_pct>=50?COL.amber:COL.neg)}">${b.win_rate_pct}%</div></div>
      <div class="sbox"><div class="l">Average gain</div><div class="v" style="color:${col(b.avg_return_pct)}">${pc(b.avg_return_pct)}</div></div>
      <div class="sbox"><div class="l">Typical (median) gain</div><div class="v" style="color:${col(b.median_pct)}">${pc(b.median_pct)}</div></div>
      <div class="sbox"><div class="l">Best case</div><div class="v" style="color:${COL.pos}">${pc(b.best_case)}</div></div>
      <div class="sbox"><div class="l">Worst case</div><div class="v" style="color:${COL.neg}">${pc(b.worst_case)}</div></div>
      <div class="sbox"><div class="l">Events tested</div><div class="v">${b.n}</div></div>
    </div>
    <div class="enote">The line shows how the stock moved on average, lined up so day 0 is the event.
      A rise before day 0 = it drifts up INTO the event; drift after 0 = it keeps re-pricing on the news.</div>
    ${t.last?`
    <div class="bhead" style="margin-top:22px">Most recent one · ${esc(t.last.date)}</div>
    <div class="chart">
      <span class="cs">move cursor over the line 👆</span>
      <span class="ct">Actual price around the last event (₹)</span>
      <canvas id="bmeter"></canvas>
    </div>
    <div class="statrow" style="grid-template-columns:repeat(3,1fr)">
      <div class="sbox"><div class="l">6 days before</div><div class="v" style="color:${COL.amber}">₹${rup(t.last.prices[0])}</div></div>
      <div class="sbox"><div class="l">On event day</div><div class="v">₹${rup(t.last.prices[t.last.days.indexOf(0)])}</div></div>
      <div class="sbox"><div class="l">3 days after</div><div class="v" style="color:${COL.pos}">₹${rup(t.last.prices[t.last.prices.length-1])}</div></div>
    </div>`:''}
    <details class="bdetails"><summary>▸ Show the full day-by-day grid (advanced)</summary>
      ${heatmap(t.grid_return,t.best,d.n_before,d.n_after)}</details>`;
  if(t.path && t.path.avg_pct.length) avgMoveChart($("bpath"),t.path.days,t.path.avg_pct,b);
  if(t.last) resultMeter($("bmeter"),t.last);
}
function avgMoveChart(cv,days,vals,best){
  const n=days.length, ev=days.indexOf(0);
  const buy=days.indexOf(-best.days_before), sell=days.indexOf(best.days_after);
  function draw(hi){
    const {ctx,w,h}=setup(cv); ctx.clearRect(0,0,w,h);
    let lo=Math.min(0,...vals), up=Math.max(0,...vals);
    const pad=((up-lo)||1)*0.18; lo-=pad; up+=pad;
    const iw=w-M.l-M.r, X=i=>M.l+(n<2?iw/2:iw*i/(n-1)), Y=v=>h-M.b-(v-lo)/((up-lo)||1)*(h-M.b-M.t);
    ctx.font="10px Consolas";ctx.textAlign="right";ctx.textBaseline="middle";
    for(let k=0;k<=4;k++){const val=lo+(up-lo)*k/4,y=Y(val);
      ctx.strokeStyle=COL.grid;ctx.beginPath();ctx.moveTo(M.l,y);ctx.lineTo(w-M.r,y);ctx.stroke();
      ctx.fillStyle=COL.mute;ctx.fillText((val>=0?"+":"")+val.toFixed(1)+"%",M.l-6,y);}
    const yz=Y(0);ctx.strokeStyle=COL.mute;ctx.lineWidth=1;ctx.beginPath();ctx.moveTo(M.l,yz);ctx.lineTo(w-M.r,yz);ctx.stroke();
    ctx.textAlign="center";ctx.textBaseline="top";ctx.fillStyle=COL.mute;
    days.forEach((d,i)=>ctx.fillText(d===0?"EVENT":(d>0?"+"+d:""+d),X(i),h-M.b+6));
    if(ev>=0){ctx.strokeStyle=COL.grid;ctx.setLineDash([4,4]);ctx.beginPath();ctx.moveTo(X(ev),M.t);ctx.lineTo(X(ev),h-M.b);ctx.stroke();ctx.setLineDash([]);}
    ctx.strokeStyle=COL.blue;ctx.lineWidth=2;ctx.beginPath();
    vals.forEach((v,i)=>{const x=X(i),y=Y(v);i?ctx.lineTo(x,y):ctx.moveTo(x,y);});ctx.stroke();
    function mark(i,c,txt){ if(i<0)return; const x=X(i),y=Y(vals[i]);
      ctx.fillStyle=c;ctx.beginPath();ctx.arc(x,y,5,0,7);ctx.fill();
      ctx.font="bold 10px Consolas";ctx.textAlign="center";ctx.textBaseline="bottom";ctx.fillText(txt,x,y-8);}
    mark(buy,COL.pos,"BUY"); mark(sell,COL.neg,"SELL");
    if(hi!=null){const x=X(hi),y=Y(vals[hi]);
      ctx.strokeStyle=COL.mute;ctx.setLineDash([3,3]);ctx.beginPath();ctx.moveTo(x,M.t);ctx.lineTo(x,h-M.b);ctx.stroke();ctx.setLineDash([]);
      ctx.fillStyle=COL.blue;ctx.beginPath();ctx.arc(x,y,5,0,7);ctx.fill();
      const dd=days[hi], lab=dd===0?"EVENT DAY":(dd>0?"T+"+dd:"T"+dd);
      const txt=lab+"   "+(vals[hi]>=0?"+":"")+vals[hi].toFixed(2)+"%";
      ctx.font="12px Consolas";const tw=ctx.measureText(txt).width+16;
      let tx=x+10; if(tx+tw>w-M.r+4)tx=x-tw-10; const ty=Math.max(M.t,y-34);
      ctx.fillStyle=COL.panel;ctx.strokeStyle=COL.blue;ctx.lineWidth=1;
      ctx.beginPath();ctx.rect(tx,ty,tw,24);ctx.fill();ctx.stroke();
      ctx.fillStyle=COL.ink;ctx.textAlign="left";ctx.textBaseline="middle";ctx.fillText(txt,tx+8,ty+12);}
  }
  draw(null);
  cv.onmousemove=e=>{const r=cv.getBoundingClientRect();const iw=r.width-M.l-M.r;
    let i=Math.round((e.clientX-r.left-M.l)/(iw/(n-1)));draw(Math.max(0,Math.min(n-1,i)));};
  cv.onmouseleave=()=>draw(null);
}
function rup(v){return Number(v).toLocaleString("en-IN",{maximumFractionDigits:2});}
function resultMeter(cv,data){
  const days=data.days, prices=data.prices, n=days.length, ev=days.indexOf(0);
  function draw(hi){
    const {ctx,w,h}=setup(cv); ctx.clearRect(0,0,w,h);
    const b=niceBounds(prices,false); if(!b)return; yaxis(ctx,w,h,b.lo,b.hi);
    const iw=w-M.l-M.r, X=i=>M.l+(n<2?iw/2:iw*i/(n-1)), Y=v=>h-M.b-(v-b.lo)/((b.hi-b.lo)||1)*(h-M.b-M.t);
    ctx.font="10px Consolas";ctx.fillStyle=COL.mute;ctx.textAlign="center";ctx.textBaseline="top";
    days.forEach((d,i)=>ctx.fillText(d===0?"EVENT":(d>0?"+"+d:""+d),X(i),h-M.b+6));
    ctx.strokeStyle=COL.blue;ctx.lineWidth=2;ctx.beginPath();
    prices.forEach((v,i)=>{const x=X(i),y=Y(v);i?ctx.lineTo(x,y):ctx.moveTo(x,y);});ctx.stroke();
    [[0,COL.amber],[ev,COL.ink],[n-1,COL.pos]].forEach(a=>{ctx.fillStyle=a[1];ctx.beginPath();ctx.arc(X(a[0]),Y(prices[a[0]]),4,0,7);ctx.fill();});
    if(hi!=null){
      const x=X(hi),y=Y(prices[hi]);
      ctx.strokeStyle=COL.mute;ctx.setLineDash([3,3]);ctx.beginPath();ctx.moveTo(x,M.t);ctx.lineTo(x,h-M.b);ctx.stroke();ctx.setLineDash([]);
      ctx.fillStyle=COL.blue;ctx.beginPath();ctx.arc(x,y,5,0,7);ctx.fill();
      const dd=days[hi], lab=dd===0?"EVENT DAY":(dd>0?"T+"+dd:"T"+dd);
      const txt=lab+"   ₹"+rup(prices[hi]);
      ctx.font="12px Consolas";const tw=ctx.measureText(txt).width+16;
      let tx=x+10; if(tx+tw>w-M.r+4)tx=x-tw-10; const ty=Math.max(M.t,y-34);
      ctx.fillStyle=COL.panel;ctx.strokeStyle=COL.blue;ctx.lineWidth=1;
      ctx.beginPath();ctx.rect(tx,ty,tw,24);ctx.fill();ctx.stroke();
      ctx.fillStyle=COL.ink;ctx.textAlign="left";ctx.textBaseline="middle";ctx.fillText(txt,tx+8,ty+12);
    }
  }
  draw(null);
  cv.onmousemove=e=>{const r=cv.getBoundingClientRect();const iw=r.width-M.l-M.r;
    let i=Math.round((e.clientX-r.left-M.l)/(iw/(n-1)));draw(Math.max(0,Math.min(n-1,i)));};
  cv.onmouseleave=()=>draw(null);
}
function verdict(win,avg){
  if(avg<=0)return['no','✕ Avoid — this historically ended in a loss'];
  if(win>=60&&avg>=0.5)return['strong','✓ Strong pattern — worth watching'];
  if(win>=55)return['ok','◐ Some edge — decent, not a sure thing'];
  return['weak','⚠ Weak — barely better than a coin flip'];
}
function pretty(et){return {RESULTS:'Quarterly Results',BOARD_MEETING:'Board Meeting',
  CORPORATE_ACTION:'Corporate Action (dividend/bonus…)',ANNOUNCEMENT:'Announcements'}[et]||et;}
function heatmap(grid,best,NB,NA){
  let mx=0; grid.forEach(r=>r.forEach(v=>{if(v!=null)mx=Math.max(mx,Math.abs(v));}));
  let h='<div class="bwrap"><table class="hm"><thead><tr><th>buy ↓ / sell →</th>';
  for(let a=1;a<=NA;a++)h+=`<th>T+${a}</th>`;
  h+='</tr></thead><tbody>';
  for(let bi=0;bi<NB;bi++){
    h+=`<tr><th>T−${bi+1}</th>`;
    for(let ai=0;ai<NA;ai++){
      const v=grid[bi][ai];
      const isB=best.days_before===bi+1&&best.days_after===ai+1;
      h+=`<td class="${isB?'bcell':''}" style="background:${cellColor(v,mx)}">${v==null?'':(v>=0?'+':'')+v}</td>`;
    }
    h+='</tr>';
  }
  h+='</tbody></table></div>';
  return h;
}
function cellColor(v,mx){
  if(v==null)return 'transparent';
  const t=Math.min(1,Math.abs(v)/(mx||1));
  return v>=0?`rgba(63,185,80,${0.10+0.55*t})`:`rgba(248,81,73,${0.10+0.55*t})`;
}

function showTab(name){
  activeTab=name;
  document.querySelectorAll(".tab").forEach(t=>t.classList.toggle("active",t.dataset.tab===name));
  const isFund=name==="fundamentals", isBeh=name==="behaviour", isRank=name==="rankings",
        isScr=name==="screener", isCmp=name==="compare", isUp=name==="upcoming";
  $("fundview").style.display=isFund?"":"none";
  $("behaviourview").style.display=isBeh?"":"none";
  $("upcomingview").style.display=isUp?"":"none";
  $("rankview").style.display=isRank?"":"none";
  $("screenview").style.display=isScr?"":"none";
  $("compareview").style.display=isCmp?"":"none";
  $("eventsview").style.display=(isFund||isBeh||isRank||isScr||isCmp||isUp)?"none":"";
  if(isFund) renderFundamentals();
  else if(isBeh) renderBehaviour();
  else if(isUp) renderUpcoming();
  else if(isRank) renderRankings();
  else if(isScr) renderScreener();
  else if(isCmp) renderCompare();
  else renderEvents(name);
}

/* ---------- Upcoming events watchlist ---------- */
let upData=null, upFilter="ALL";
async function renderUpcoming(){
  const uv=$("upcomingview");
  if(!upData){
    uv.innerHTML='<div class="enote">Finding events still ahead across all stocks…</div>';
    try{ upData=await (await fetch("/api/upcoming")).json(); }
    catch(e){ uv.innerHTML='<div class="banner">Couldn’t load upcoming events: '+esc(e)+'</div>'; return; }
  }
  if(activeTab!=="upcoming")return;
  if(upData.error){ uv.innerHTML='<div class="banner">'+esc(upData.error)+'</div>'; return; }
  drawUpcoming();
}
function setUpFilter(f){ upFilter=f; drawUpcoming(); }
function drawUpcoming(){
  const uv=$("upcomingview"), d=upData; if(!d||!d.rows)return;
  const evs=["ALL","RESULTS","BOARD_MEETING","CORPORATE_ACTION"];
  const tabs=evs.map(k=>`<button class="rtab${k===upFilter?' active':''}" onclick="setUpFilter('${k}')">${k==="ALL"?"All":esc(pretty(k))}</button>`).join('');
  let rows=d.rows.filter(r=>upFilter==="ALL"||r.event===upFilter);
  const body=rows.map(r=>{
    const wc=r.win==null?'var(--mute)':(r.win>=55?COL.pos:(r.win>=50?COL.amber:COL.neg));
    const when=r.in_days<=0?'today':(r.in_days===1?'tomorrow':'in '+r.in_days+'d');
    const badge=r.status==='confirmed'?'<span class="ubadge conf">confirmed</span>':'<span class="ubadge est">estimated</span>';
    return `<tr class="rrow" onclick="goSym('${r.symbol}')" title="open ${esc(r.symbol)}">
      <td class="rnum" style="white-space:nowrap">${esc(r.date)}<div class="uwhen">${when}</div></td>
      <td class="rsym">${esc(r.symbol)}</td>
      <td>${esc(pretty(r.event))} ${badge}</td>
      <td class="rnum">${r.db!=null?`buy ${r.db}d / sell ${r.da}d`:'—'}</td>
      <td class="rnum" style="color:${wc};font-weight:700">${r.win==null?'—':r.win+'%'}</td>
      <td class="rnum">${r.events!=null?r.events:'—'}</td></tr>`;
  }).join('');
  uv.innerHTML=`
    <div class="bintro">📅 <b>Your watchlist</b> — corporate events still ahead (next ${d.horizon_days} days), each with the stock’s
      historical <b>best rule</b> and its win rate. <b>Confirmed</b> = the exact date is already announced;
      <b>estimated</b> = projected from the stock’s past cadence. Click a row to open the stock.</div>
    <div class="rtabs">${tabs}</div>
    <div class="tblwrap" style="margin-top:12px"><table>
      <thead><tr><th>Date</th><th>Stock</th><th>Event</th><th style="text-align:right">Best rule</th>
      <th style="text-align:right">Win rate</th>
      <th style="text-align:right">Events</th></tr></thead>
      <tbody>${body||'<tr><td colspan="6" class="enote">no upcoming events in this window</td></tr>'}</tbody></table></div>
    <div class="enote">${rows.length} upcoming event(s). Dates are read from NSE filings (confirmed) or projected from
      cadence (estimated). Win rate = historical best buy/sell rule for that event.</div>`;
}

/* ---------- Rankings: all stocks ordered by win rate, per event ---------- */
let rankEvent="RESULTS", rankData=null;
async function renderRankings(){
  const rv=$("rankview");
  if(!rankData){
    rv.innerHTML='<div class="enote">Ranking every stock by win rate — this analyses all of them, first load takes a few seconds…</div>';
    try{ rankData=await (await fetch("/api/rankings")).json(); }
    catch(e){ rv.innerHTML='<div class="banner">Couldn’t build rankings: '+esc(e)+'</div>'; return; }
  }
  if(activeTab!=="rankings")return;
  if(rankData.error){ rv.innerHTML='<div class="banner">'+esc(rankData.error)+'</div>'; return; }
  drawRankings();
}
function setRankEvent(et){ rankEvent=et; drawRankings(); }
function drawRankings(){
  const rv=$("rankview"), d=rankData; if(!d||!d.ranks)return;
  const subs=(d.events||["RESULTS","BOARD_MEETING","CORPORATE_ACTION","ANNOUNCEMENT"]);
  const tabs=subs.map(k=>`<button class="rtab${k===rankEvent?' active':''}" onclick="setRankEvent('${k}')">${esc(pretty(k))}</button>`).join('');
  const list=d.ranks[rankEvent]||[];
  const rows=list.map((r,i)=>{
    const barcol=r.win>=55?COL.pos:(r.win>=50?COL.amber:COL.neg);
    const medal=i===0?'🥇':i===1?'🥈':i===2?'🥉':(i+1);
    return `<tr class="rrow" onclick="goSym('${r.symbol}')" title="open ${esc(r.symbol)}">
      <td class="rk">${medal}</td>
      <td class="rsym">${esc(r.symbol)}</td>
      <td><div class="rbar"><span style="width:${Math.max(2,Math.min(100,r.win))}%;background:${barcol}"></span></div></td>
      <td class="rnum" style="color:${barcol};font-weight:700">${r.win}%</td>
      <td class="rnum" style="color:${r.avg>=0?COL.pos:COL.neg}">${r.avg>=0?'+':''}${r.avg}%</td>
      <td class="rnum">buy ${r.db}d / sell ${r.da}d</td>
      <td class="rnum">${r.events}</td></tr>`;
  }).join('');
  rv.innerHTML=`
    <div class="bintro">All <b>${d.n_stocks}</b> tracked stocks — the Nifty 50 plus a couple of recent
      ex-members kept for their history — ranked by <b>win rate</b> for each event, highest first.
      Click a sub-tab to reorder; click a stock to open its Behaviour view.</div>
    <div class="rtabs">${tabs}</div>
    <div class="tblwrap" style="margin-top:12px"><table>
      <thead><tr><th>#</th><th>Stock</th><th>Win rate</th><th style="text-align:right">Win %</th>
      <th style="text-align:right">Avg gain</th><th style="text-align:right">Best rule</th>
      <th style="text-align:right">Events</th></tr></thead>
      <tbody>${rows||'<tr><td colspan="7" class="enote">no stocks with this event yet</td></tr>'}</tbody></table></div>
    <div class="enote">${list.length} of ${d.n_stocks} stocks have ${esc(pretty(rankEvent))} data.
      Win rate = share of past events where the best buy/sell rule made money.</div>`;
}
function goSym(sym){ const i=SYMS.indexOf(sym); if(i<0)return; idx=i; loadStock(); showTab('behaviour'); }

/* ---------- Screener: filter all stocks by fundamentals ---------- */
let screenData=null, screenSort={key:'roce',dir:-1};
const SCOLS=[
  {k:'symbol',t:'Stock',fmt:v=>v,align:'left'},
  {k:'pe',t:'P/E',fmt:v=>v==null?'—':v.toFixed(1),good:'low'},
  {k:'roce',t:'ROCE %',fmt:v=>v==null?'—':v.toFixed(0)+'%',good:'high'},
  {k:'opm',t:'OPM %',fmt:v=>v==null?'—':v.toFixed(0)+'%',good:'high'},
  {k:'sales_g',t:'Sales gr%',fmt:v=>v==null?'—':(v>=0?'+':'')+v.toFixed(0)+'%',good:'high'},
  {k:'profit_g',t:'Profit gr%',fmt:v=>v==null?'—':(v>=0?'+':'')+v.toFixed(0)+'%',good:'high'},
  {k:'de',t:'Debt/Eq',fmt:v=>v==null?'—':v.toFixed(2),good:'low'},
  {k:'npat',t:'Net Profit ₹cr',fmt:v=>v==null?'—':Math.round(v).toLocaleString(),align:'right'},
];
async function renderScreener(){
  const sv=$("screenview");
  if(!screenData){
    sv.innerHTML='<div class="enote">Loading fundamentals for every stock…</div>';
    try{ screenData=await (await fetch("/api/screener")).json(); }
    catch(e){ sv.innerHTML='<div class="banner">Couldn’t load screener: '+esc(e)+'</div>'; return; }
  }
  if(activeTab!=="screener")return;
  if(screenData.error){ sv.innerHTML='<div class="banner">'+esc(screenData.error)+'</div>'; return; }
  if(!screenData.stocks||!screenData.stocks.length){
    sv.innerHTML='<div class="banner">No fundamentals loaded yet. Hard-refresh the page (Ctrl+Shift+R).</div>'; return; }
  // Build the filter shell ONCE; only #sctable redraws on change so the inputs
  // keep their values and never lose focus while you type.
  sv.innerHTML=`
    <div class="bintro">Filter the tracked universe (<b>${screenData.stocks.length}</b> stocks — the Nifty 50 plus a couple
      of recent index changes) by fundamentals to find <b>high-quality &amp; cheap</b> names.
      Blank = no limit. Click a column to sort; click a stock to open it.</div>
    <div class="scfilters">
      <label>ROCE ≥ <input id="f_roce" type="number" placeholder="any" oninput="drawScreener()"></label>
      <label>Profit gr% ≥ <input id="f_pg" type="number" placeholder="any" oninput="drawScreener()"></label>
      <label>OPM% ≥ <input id="f_opm" type="number" placeholder="any" oninput="drawScreener()"></label>
      <label>Debt/Eq ≤ <input id="f_de" type="number" step="0.1" placeholder="any" oninput="drawScreener()"></label>
      <label>P/E ≤ <input id="f_pe" type="number" placeholder="any" oninput="drawScreener()"></label>
      <button class="scbtn" onclick="scPreset()">✨ High-quality + cheap</button>
      <button class="scbtn ghost" onclick="scClear()">Clear</button>
    </div>
    <div id="sctable"></div>`;
  drawScreener();
}
function scGet(id){ const v=parseFloat(($(id)||{}).value); return isNaN(v)?null:v; }
function scPreset(){ $("f_roce").value=15; $("f_pg").value=10; $("f_de").value=0.5; $("f_pe").value=25; drawScreener(); }
function scClear(){ ["f_roce","f_pg","f_de","f_pe","f_opm"].forEach(i=>{if($(i))$(i).value='';}); drawScreener(); }
function scSort(k){ if(screenSort.key===k)screenSort.dir*=-1; else screenSort={key:k,dir:(k==='symbol'||k==='pe'||k==='de')?1:-1}; drawScreener(); }
function drawScreener(){
  const tbl=$("sctable"); if(!tbl||!screenData||!screenData.stocks)return;
  const all=screenData.stocks;
  const minR=scGet('f_roce'), minPG=scGet('f_pg'), maxDE=scGet('f_de'), maxPE=scGet('f_pe'), minOPM=scGet('f_opm');
  let rows=all.filter(s=>
    (minR==null||(s.roce!=null&&s.roce>=minR)) &&
    (minPG==null||(s.profit_g!=null&&s.profit_g>=minPG)) &&
    (maxDE==null||(s.de!=null&&s.de<=maxDE)) &&
    (maxPE==null||(s.pe!=null&&s.pe<=maxPE)) &&
    (minOPM==null||(s.opm!=null&&s.opm>=minOPM)));
  const k=screenSort.key, dir=screenSort.dir;
  rows.sort((a,b)=>{ let x=a[k],y=b[k];
    if(x==null)return 1; if(y==null)return -1;
    if(k==='symbol')return x<y?-dir:x>y?dir:0; return (x-y)*dir; });
  const colcol=(c,v)=>{ if(v==null||!c.good)return '';
    const vals=rows.map(r=>r[c.k]).filter(x=>x!=null); if(!vals.length)return '';
    const mn=Math.min(...vals),mx=Math.max(...vals); if(mn===mx)return '';
    const t=(v-mn)/(mx-mn), q=c.good==='high'?t:1-t;
    return q>=0.66?'style="color:'+COL.pos+'"':q<=0.33?'style="color:'+COL.neg+'"':''; };
  const head=SCOLS.map(c=>`<th class="scH ${c.align==='left'?'':'rt'}" onclick="scSort('${c.k}')">${esc(c.t)}${screenSort.key===c.k?(dir>0?' ▲':' ▼'):''}</th>`).join('');
  const body=rows.map(s=>'<tr class="rrow" onclick="goSym(\''+s.symbol+'\')">'+
    SCOLS.map(c=>`<td class="${c.align==='left'?'sym':'rnum'}" ${colcol(c,s[c.k])}>${c.fmt(s[c.k])}</td>`).join('')+'</tr>').join('');
  tbl.innerHTML=`
    <div class="tblwrap" style="margin-top:10px"><table><thead><tr>${head}</tr></thead><tbody>${body||'<tr><td colspan="8" class="enote">No stock matches these filters.</td></tr>'}</tbody></table></div>
    <div class="enote">${rows.length} of ${all.length} stocks match. Green = better vs peers, red = worse. Latest annual figures; P/E = price ÷ latest EPS.</div>`;
}

/* ---------- Compare: 2-3 stocks side by side ---------- */
let cmpSel=[], cmpBeh={};
async function renderCompare(){
  const cv=$("compareview");
  if(!screenData){ try{ screenData=await (await fetch("/api/screener")).json(); }catch(e){} }
  if(activeTab!=="compare")return;
  if(!cmpSel.length){ cmpSel=[SYMS[idx]].filter(Boolean); }
  drawCompare();
}
function cmpAdd(sym){ sym=(sym||'').toUpperCase(); if(!sym||cmpSel.includes(sym)||!SYMS.includes(sym))return;
  if(cmpSel.length>=3)cmpSel.shift(); cmpSel.push(sym); $("cmppick").value=''; drawCompare(); }
function cmpDel(sym){ cmpSel=cmpSel.filter(s=>s!==sym); drawCompare(); }
async function cmpLoadBeh(sym){ if(cmpBeh[sym])return cmpBeh[sym];
  try{ cmpBeh[sym]=await (await fetch("/api/behaviour?sym="+encodeURIComponent(sym))).json(); }catch(e){ cmpBeh[sym]={}; }
  return cmpBeh[sym]; }
async function drawCompare(){
  const cv=$("compareview");
  const map={}; (screenData&&screenData.stocks||[]).forEach(s=>map[s.symbol]=s);
  await Promise.all(cmpSel.map(cmpLoadBeh));
  if(activeTab!=="compare")return;
  const chips=cmpSel.map(s=>`<span class="cmpchip">${esc(s)} <b onclick="cmpDel('${s}')">✕</b></span>`).join('');
  const FR=[['price','Price','₹'],['pe','P/E',''],['roce','ROCE','%'],['opm','OPM','%'],
    ['sales_g','Sales growth','%'],['profit_g','Profit growth','%'],['de','Debt/Equity',''],
    ['sales','Sales','₹cr'],['npat','Net Profit','₹cr'],['eps','EPS','₹']];
  const goodHigh=new Set(['roce','opm','sales_g','profit_g','npat','sales','eps']);
  const goodLow=new Set(['pe','de']);
  const fmt=(k,v)=>v==null?'—':(k==='sales'||k==='npat'?Math.round(v).toLocaleString():(k==='price'||k==='eps'?v.toLocaleString(undefined,{maximumFractionDigits:2}):v))+(FR.find(r=>r[0]===k)[2]==='%'?'%':'');
  const frows=FR.map(([k,lab])=>{
    const vals=cmpSel.map(s=>map[s]?map[s][k]:null).filter(v=>v!=null);
    const best=goodHigh.has(k)?Math.max(...vals):goodLow.has(k)?Math.min(...vals):null;
    return `<tr><td class="cmplab">${esc(lab)}</td>`+cmpSel.map(s=>{ const v=map[s]?map[s][k]:null;
      const hit=(best!=null&&v===best&&vals.length>1);
      return `<td class="rnum"${hit?' style="color:'+COL.pos+';font-weight:700"':''}>${fmt(k,v)}</td>`;}).join('')+'</tr>';
  }).join('');
  const EVT=[['RESULTS','Results'],['BOARD_MEETING','Board Mtg'],['CORPORATE_ACTION','Corp Action'],['ANNOUNCEMENT','Announce']];
  const brows=EVT.map(([et,lab])=>`<tr><td class="cmplab">${esc(lab)} — rule · win</td>`+
    cmpSel.map(s=>{ const t=(cmpBeh[s]&&cmpBeh[s].types||{})[et];
      if(!t||!t.best)return '<td class="rnum muted">—</td>';
      const b=t.best,wc=b.win_rate_pct>=55?COL.pos:(b.win_rate_pct>=50?COL.amber:COL.neg);
      return `<td class="rnum"><span class="mono">buy${b.days_before}/sell${b.days_after}</span><br><b style="color:${wc}">${b.win_rate_pct}%</b> · ${b.avg_return_pct>=0?'+':''}${b.avg_return_pct}%</td>`;
    }).join('')+'</tr>').join('');
  const hdr='<th></th>'+cmpSel.map(s=>`<th class="rt"><span class="cmphead" onclick="goSym('${s}')">${esc(s)}</span></th>`).join('');
  cv.innerHTML=`
    <div class="bintro">Put <b>2–3 stocks side by side</b> — fundamentals and event behaviour. Best value in each row is highlighted.</div>
    <div class="cmpbar">${chips}
      <input id="cmppick" list="cmplist" placeholder="+ add stock" onchange="cmpAdd(this.value)">
      <datalist id="cmplist">${SYMS.map(s=>'<option value="'+esc(s)+'">').join('')}</datalist></div>
    ${cmpSel.length?`<div class="tblwrap" style="margin-top:10px"><table>
      <thead><tr>${hdr}</tr></thead>
      <tbody><tr class="cmpsec"><td colspan="${cmpSel.length+1}">Fundamentals (latest year)</td></tr>${frows}
      <tr class="cmpsec"><td colspan="${cmpSel.length+1}">Event behaviour (best rule)</td></tr>${brows}</tbody></table></div>`
      :'<div class="enote">Add a stock to compare.</div>'}`;
}

/* ---------- navigation ---------- */
function go(i){ idx=Math.max(0,Math.min(i,SYMS.length-1)); hideAC(); loadStock(); }

/* ---------- live search dropdown ---------- */
const AC=$("ac"), SEARCH=$("search");
let acItems=[], acSel=-1;
function hideAC(){ AC.style.display="none"; acSel=-1; }
function renderAC(){
  const q=SEARCH.value.trim().toUpperCase();
  if(!q){ hideAC(); return; }
  const starts=[], incl=[];
  for(const s of SYMS){
    if(s.startsWith(q)) starts.push(s);
    else if(s.includes(q)) incl.push(s);
  }
  const total=starts.length+incl.length;
  acItems=starts.concat(incl).slice(0,30);
  acSel=-1;
  if(!acItems.length){ AC.innerHTML='<div class="acempty">no match for “'+esc(q)+'”</div>'; AC.style.display="block"; return; }
  let html=acItems.map((s,i)=>{
    const k=s.indexOf(q);
    const hl=k<0?esc(s):esc(s.slice(0,k))+'<b>'+esc(s.slice(k,k+q.length))+'</b>'+esc(s.slice(k+q.length));
    const tag=READY.has(s)?'<span class="rdy">✓ data</span>':'';
    return `<div class="acitem" data-i="${i}"><span>${hl}</span>${tag}</div>`;
  }).join('');
  if(total>acItems.length) html+=`<div class="acmore">${total.toLocaleString()} matches — showing first ${acItems.length}</div>`;
  AC.innerHTML=html; AC.style.display="block";
}
function pickAC(i){
  if(i<0||i>=acItems.length)return;
  const j=SYMS.indexOf(acItems[i]);
  if(j>=0){ SEARCH.value=acItems[i]; go(j); SEARCH.blur(); }
}
function moveAC(d){
  if(AC.style.display==="none"||!acItems.length)return;
  acSel=(acSel+d+acItems.length)%acItems.length;
  [...AC.querySelectorAll(".acitem")].forEach((el,i)=>el.classList.toggle("sel",i===acSel));
  const el=AC.querySelector(".acitem.sel"); if(el)el.scrollIntoView({block:"nearest"});
}
SEARCH.addEventListener("input",renderAC);
SEARCH.addEventListener("focus",()=>{ if(SEARCH.value.trim())renderAC(); });
SEARCH.addEventListener("blur",()=>setTimeout(hideAC,150));   // allow click first
SEARCH.addEventListener("keydown",e=>{
  if(e.key==="ArrowDown"){ e.preventDefault(); moveAC(1); }
  else if(e.key==="ArrowUp"){ e.preventDefault(); moveAC(-1); }
  else if(e.key==="Enter"){
    e.preventDefault();
    if(acSel>=0) pickAC(acSel);
    else if(acItems.length) pickAC(0);            // best match
  }
  else if(e.key==="Escape"){ hideAC(); }
});
AC.addEventListener("mousedown",e=>{
  const it=e.target.closest(".acitem"); if(it) pickAC(+it.dataset.i);
});

$("first").onclick=()=>go(0);
$("prev").onclick=()=>go(idx-1);
$("next").onclick=()=>go(idx+1);
$("last").onclick=()=>go(SYMS.length-1);
document.querySelectorAll(".tab").forEach(t=>t.onclick=()=>showTab(t.dataset.tab));
document.addEventListener("keydown",e=>{
  if(document.activeElement.id==="search")return;
  if(e.key==="ArrowLeft")go(idx-1);
  else if(e.key==="ArrowRight")go(idx+1);
  else if(e.key==="Home")go(0);
  else if(e.key==="End")go(SYMS.length-1);
});
window.addEventListener("resize",()=>{ if(activeTab==="fundamentals"&&stock)renderFundamentals(); });

(async ()=>{
  const s=await (await fetch(SYMBOLS_URL)).json();
  SYMS=s.symbols;
  READY=new Set(s.ready||[]);
  go(0);                       // start from the FIRST stock
})();
</script>
</body>
</html>"""


# ---------------------------------------------------------------------------
# HTTP handler
# ---------------------------------------------------------------------------
SYMBOLS = discover_symbols()


class Handler(BaseHTTPRequestHandler):
    def _send(self, code, body, ctype):
        data = body.encode("utf-8") if isinstance(body, str) else body
        self.send_response(code)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-store")      # always serve fresh
        self.end_headers()
        self.wfile.write(data)

    def _json(self, obj, code=200):
        self._send(code, json.dumps(obj), "application/json; charset=utf-8")

    def _send_file(self, data: bytes, ctype: str, filename: str, inline: bool = False):
        how = "inline" if inline else "attachment"
        self.send_response(200)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Content-Disposition", f'{how}; filename="{filename}"')
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(data)

    def do_GET(self):
        parsed = urlparse(self.path)
        route = parsed.path
        qs = parse_qs(parsed.query)

        if route in ("/", "/index.html"):
            self._send(200, PAGE, "text/html; charset=utf-8")
        elif route == "/api/symbols":
            self._json({"symbols": SYMBOLS, "count": len(SYMBOLS),
                        "ready": downloaded_symbols()})
        elif route == "/api/file":                       # proxy an NSE attachment
            url = (qs.get("url", [""])[0]).strip()
            raw = qs.get("raw", ["0"])[0] == "1"
            dl = qs.get("dl", ["0"])[0] == "1"           # force a download
            if not url.startswith("http"):
                self._send(400, "bad file url", "text/plain; charset=utf-8")
                return
            try:
                data, ctype, name = fetch_nse_file(url)
                # XBRL/XML attachments -> render a readable page, not raw XML
                if name.lower().endswith(".xml") and not raw:
                    try:
                        self._send(200, xml_to_html(data, name, url),
                                   "text/html; charset=utf-8")
                        return
                    except Exception:                    # noqa: BLE001 - fall back to raw
                        pass
                self._send_file(data, ctype, name, inline=not dl)
            except Exception as e:                       # noqa: BLE001
                self._send(502, f"Could not fetch the file from NSE.\n\n"
                                f"{type(e).__name__}: {e}", "text/plain; charset=utf-8")

        elif route == "/api/xbrl_excel":
            sym = (qs.get("sym", [""])[0]).strip().upper()
            url = (qs.get("url", [""])[0]).strip()
            if not url.startswith("http"):
                self._send(400, "bad xbrl url", "text/plain; charset=utf-8")
                return
            try:
                data, fname = build_xbrl_excel(sym, url)
                self._send_file(
                    data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    fname)
            except Exception as e:                       # noqa: BLE001
                self._send(500, f"Could not build the Excel file.\n\n{type(e).__name__}: {e}",
                           "text/plain; charset=utf-8")

        elif route == "/api/statements_excel":
            # financial statements straight from the local folders (offline)
            sym = (qs.get("sym", [""])[0]).strip().upper()
            try:
                data, fname = build_statements_excel(sym)
                self._send_file(
                    data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    fname)
            except Exception as e:                       # noqa: BLE001
                self._send(404, f"Could not build the Excel file.\n\n{type(e).__name__}: {e}",
                           "text/plain; charset=utf-8")

        elif route == "/api/rankings":
            try:
                self._json(build_rankings())
            except Exception as e:                       # noqa: BLE001
                self._json({"error": f"{type(e).__name__}: {e}"}, code=500)

        elif route == "/api/screener":
            try:
                self._json(build_screener())
            except Exception as e:                       # noqa: BLE001
                self._json({"error": f"{type(e).__name__}: {e}"}, code=500)

        elif route == "/api/upcoming":
            try:
                self._json(build_upcoming())
            except Exception as e:                       # noqa: BLE001
                self._json({"error": f"{type(e).__name__}: {e}"}, code=500)

        elif route in ("/api/stock", "/api/events", "/api/fetch", "/api/behaviour"):
            sym = (qs.get("sym", [""])[0]).strip().upper()
            if sym not in SYMBOLS:
                self._json({"error": f"unknown symbol '{sym}'"}, code=404)
                return
            try:
                if route == "/api/stock":
                    self._json(build_payload(sym))
                elif route == "/api/events":
                    self._json(build_events(sym))
                elif route == "/api/behaviour":
                    self._json(build_behaviour(sym))
                else:                                    # /api/fetch -> download NSE
                    res = fetch_nse_feeds(sym)
                    self._json({"ok": res["any"], "symbol": sym,
                                "counts": res["counts"], "errors": res["errors"]})
            except Exception as e:                       # noqa: BLE001
                self._json({"error": f"{type(e).__name__}: {e}"}, code=500)
        else:
            self._send(404, "not found", "text/plain; charset=utf-8")

    def log_message(self, *args):
        pass


def lan_ip() -> str:
    """This machine's LAN IP (for --lan network access)."""
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(("8.8.8.8", 80))
        return s.getsockname()[0]
    except Exception:                              # noqa: BLE001
        return "127.0.0.1"
    finally:
        s.close()


def main():
    port = next((int(a) for a in sys.argv[1:] if a.isdigit()), 8000)
    lan = "--lan" in sys.argv
    if not SYMBOLS:
        raise SystemExit("No stock fundamentals found in pnl/ quarterly/ ... folders.")

    host = "0.0.0.0" if lan else "127.0.0.1"
    ip = lan_ip() if lan else "localhost"
    url = f"http://{ip}:{port}"
    server = ThreadingHTTPServer((host, port), Handler)
    print("=" * 62)
    print(f"  NSE Stock Browser  —  serving {len(SYMBOLS):,} stocks")
    print(f"  Open:  {url}" + ("   (reachable from other devices on your LAN)" if lan else ""))
    print(f"  First: {SYMBOLS[0]}   Last: {SYMBOLS[-1]}")
    print("  Press Ctrl+C to stop.")
    print("=" * 62)

    if "--no-open" not in sys.argv:
        threading.Timer(0.6, lambda: webbrowser.open(f"http://localhost:{port}")).start()
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nStopped.")
        server.shutdown()


if __name__ == "__main__":
    main()
